"""Tool: Multi-investor LP DDQ exporter.

Generates Due Diligence Questionnaire responses in various LP formats
by combining data from the impact engine, framework assessments, and DD checklist.
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field

from openharness.impact.database import get_metric_store
from openharness.impact.five_dimensions import assess_five_dimensions
from openharness.impact.gap_analysis import analyze_gaps
from openharness.impact.models import Company
from openharness.impact.sdg_mapper import map_sdg_alignment
from openharness.tools.impact.common import infer_themes, normalize_metric_map, normalize_sdg_goals
from openharness.tools.base import BaseTool, ToolExecutionContext, ToolResult


LP_DDQ_TEMPLATES = {
    "ilpa": {
        "name": "ILPA DDQ (ESG Section)",
        "description": "Institutional Limited Partners Association standardized DDQ, Section 10: ESG",
        "sections": [
            {"id": "ilpa-10.1", "question": "Describe the firm's approach to integrating ESG considerations into the investment process.", "data_sources": ["impact_thesis", "sdg_alignment", "five_dimensions"]},
            {"id": "ilpa-10.2", "question": "Does the firm have a responsible investment or ESG policy? If so, please provide a copy.", "data_sources": ["governance"]},
            {"id": "ilpa-10.3", "question": "Has the firm committed to any responsible investment frameworks or principles (e.g., UNPRI)?", "data_sources": ["unpri"]},
            {"id": "ilpa-10.4", "question": "Describe how ESG risks are identified, assessed, and managed in the due diligence process.", "data_sources": ["dd_checklist", "risk"]},
            {"id": "ilpa-10.5", "question": "Describe how the firm monitors and manages ESG risks during the investment holding period.", "data_sources": ["measurement_systems"]},
            {"id": "ilpa-10.6", "question": "Describe any ESG-related training provided to investment professionals.", "data_sources": ["team"]},
            {"id": "ilpa-10.7", "question": "Does the firm measure or report on the impact of its investments? If so, what metrics or frameworks are used?", "data_sources": ["iris_metrics", "sdg_alignment", "edci"]},
            {"id": "ilpa-10.8", "question": "Has the firm experienced any material ESG incidents in its portfolio? If so, describe the response.", "data_sources": ["risk"]},
        ],
    },
    "giin_iris": {
        "name": "GIIN / IRIS+ Impact Report",
        "description": "Standard impact reporting template using IRIS+ metrics and 5 Dimensions",
        "sections": [
            {"id": "giin-1", "question": "Fund/Company overview and impact thesis", "data_sources": ["impact_thesis"]},
            {"id": "giin-2", "question": "Impact goals and SDG alignment", "data_sources": ["sdg_alignment"]},
            {"id": "giin-3", "question": "5 Dimensions of Impact assessment", "data_sources": ["five_dimensions"]},
            {"id": "giin-4", "question": "IRIS+ Core Metric Set reporting", "data_sources": ["iris_metrics", "gap_analysis"]},
            {"id": "giin-5", "question": "Impact performance and outcomes", "data_sources": ["outcomes"]},
            {"id": "giin-6", "question": "Impact measurement methodology", "data_sources": ["measurement_systems"]},
            {"id": "giin-7", "question": "Impact risks and mitigation", "data_sources": ["risk"]},
        ],
    },
    "edci": {
        "name": "EDCI Annual Survey",
        "description": "ESG Data Convergence Initiative annual portfolio company survey",
        "sections": [
            {"id": "edci-e", "question": "Environment metrics (Scope 1/2/3, renewable energy, net zero)", "data_sources": ["edci_environment"]},
            {"id": "edci-s", "question": "Social metrics (injuries, diversity, engagement, wages)", "data_sources": ["edci_social"]},
            {"id": "edci-g", "question": "Governance metrics (board independence, data privacy, ESG oversight)", "data_sources": ["edci_governance"]},
        ],
    },
    "sfdr": {
        "name": "SFDR Annex III/IV Disclosure",
        "description": "EU SFDR pre-contractual and periodic disclosure template (Article 8/9 products)",
        "sections": [
            {"id": "sfdr-class", "question": "Fund classification under SFDR (Article 6/8/9) and rationale", "data_sources": ["sfdr_classification"]},
            {"id": "sfdr-esg", "question": "Environmental or social characteristics promoted (Art.8) / sustainable investment objective (Art.9)", "data_sources": ["impact_thesis", "sdg_alignment"]},
            {"id": "sfdr-pai", "question": "Principal Adverse Impact indicators — 14 mandatory + relevant optional", "data_sources": ["sfdr_pai"]},
            {"id": "sfdr-dnsh", "question": "Do No Significant Harm assessment for sustainable investments", "data_sources": ["risk", "exclusion_screening"]},
            {"id": "sfdr-taxonomy", "question": "Proportion of EU Taxonomy-aligned investments", "data_sources": ["taxonomy_alignment"]},
            {"id": "sfdr-data", "question": "Data sources and limitations", "data_sources": ["measurement_systems"]},
            {"id": "sfdr-entity", "question": "Entity-level sustainability risk integration (Art. 3-5)", "data_sources": ["governance"]},
        ],
    },
    "custom": {
        "name": "Custom LP DDQ",
        "description": "Flexible template combining data from all available frameworks",
        "sections": [
            {"id": "custom-1", "question": "Impact thesis and SDG contribution", "data_sources": ["impact_thesis", "sdg_alignment"]},
            {"id": "custom-2", "question": "Impact measurement framework and metrics", "data_sources": ["iris_metrics", "five_dimensions"]},
            {"id": "custom-3", "question": "ESG risk assessment", "data_sources": ["risk", "sfdr_pai"]},
            {"id": "custom-4", "question": "Climate and environmental performance", "data_sources": ["tcfd", "edci_environment"]},
            {"id": "custom-5", "question": "Social and governance performance", "data_sources": ["edci_social", "edci_governance"]},
            {"id": "custom-6", "question": "Gap analysis and improvement plan", "data_sources": ["gap_analysis", "dd_checklist"]},
        ],
    },
}


class LpDdqExportInput(BaseModel):
    template: Literal["ilpa", "giin_iris", "edci", "sfdr", "custom"] = Field(
        description="LP DDQ template to use: 'ilpa' (ILPA ESG section), 'giin_iris' (GIIN/IRIS+ report), 'edci' (EDCI survey), 'sfdr' (SFDR Annex III/IV), 'custom' (all frameworks)",
    )
    action: Literal["list_templates", "generate", "preview"] = Field(
        default="generate",
        description="'list_templates': Show available templates. 'generate': Create DDQ response. 'preview': Show template structure.",
    )
    company_name: str = Field(default="", description="Company name")
    company_description: str = Field(default="", description="Company description")
    sector: str = Field(default="", description="Industry sector")
    geography: str = Field(default="", description="Country or region (e.g. 'Kenya', 'Southeast Asia')")
    stage: Literal["", "pre-seed", "seed", "series-a", "series-b", "growth", "mature"] = Field(
        default="", description="Investment stage",
    )
    impact_themes: list[str] = Field(default_factory=list, description="Impact themes")
    sdg_claims: list[int] = Field(default_factory=list, description="Claimed SDGs (1-17)")
    reported_metrics: dict[str, str] = Field(default_factory=dict, description="IRIS+ metric ID -> value")
    output_format: Literal["text", "json", "csv", "xlsx"] = Field(default="text", description="Output format ('xlsx' for Excel)")
    output_path: str = Field(default="", description="File path to save the output (required for xlsx)")
    narrative_mode: Literal["data", "narrative_prompt"] = Field(
        default="data",
        description=(
            "'data': Standard data-driven output. "
            "'narrative_prompt': Output includes structured LLM prompts for each section "
            "so the agent can generate polished prose narratives."
        ),
    )
    draft_review: bool = Field(
        default=False,
        description="If True, output is wrapped with DRAFT markers for human review before finalization.",
    )


class LpDdqExportTool(BaseTool):
    name = "lp_ddq_export"
    description = (
        "Generate LP Due Diligence Questionnaire responses in standard formats. "
        "Templates: ILPA (ESG section), GIIN/IRIS+ impact report, EDCI annual survey, "
        "or custom multi-framework. Combines data from IRIS+ catalog, SDG alignment, "
        "5 Dimensions, gap analysis, and ESG frameworks into structured DDQ responses."
    )
    input_model = LpDdqExportInput

    def is_read_only(self, arguments: BaseModel) -> bool:
        args = arguments if isinstance(arguments, LpDdqExportInput) else LpDdqExportInput.model_validate(arguments)
        if args.output_format == "xlsx":
            return False
        return not bool(args.output_path)

    async def execute(self, arguments: BaseModel, context: ToolExecutionContext) -> ToolResult:
        args = arguments if isinstance(arguments, LpDdqExportInput) else LpDdqExportInput.model_validate(arguments)

        if args.action == "list_templates":
            return self._list_templates()

        template = LP_DDQ_TEMPLATES.get(args.template)
        if not template:
            return ToolResult(output=f"Unknown template: {args.template}", is_error=True)

        if args.action == "preview":
            return self._preview(template)

        if args.action == "generate":
            if not args.company_name:
                return ToolResult(output="company_name is required for generate", is_error=True)
            return self._generate(args, template, context)

        return ToolResult(output=f"Unknown action: {args.action}", is_error=True)

    def _list_templates(self) -> ToolResult:
        lines = ["Available LP DDQ Templates:\n"]
        for key, tmpl in LP_DDQ_TEMPLATES.items():
            lines.append(f"  {key}: {tmpl['name']}")
            lines.append(f"    {tmpl['description']}")
            lines.append(f"    Sections: {len(tmpl['sections'])}")
            lines.append("")
        return ToolResult(output="\n".join(lines))

    def _preview(self, template: dict) -> ToolResult:
        lines = [f"Template: {template['name']}\n{template['description']}\n"]
        for section in template["sections"]:
            lines.append(f"  [{section['id']}] {section['question']}")
            lines.append(f"    Data sources: {', '.join(section['data_sources'])}")
        return ToolResult(output="\n".join(lines))

    def _generate(self, args: LpDdqExportInput, template: dict, context: ToolExecutionContext | None = None) -> ToolResult:
        try:
            store = get_metric_store()
        except FileNotFoundError as e:
            return ToolResult(output=str(e), is_error=True)

        reported_metrics, _ = normalize_metric_map(args.reported_metrics)
        sdg_claims, _ = normalize_sdg_goals(args.sdg_claims)
        company = Company(
            name=args.company_name,
            description=args.company_description,
            sector=args.sector,
            geography=args.geography,
            stage=args.stage,
            impact_themes=infer_themes(f"{args.company_description} {args.sector}", args.impact_themes),
            reported_metrics=reported_metrics,
            sdg_claims=sdg_claims,
        )

        data_cache: dict = {}

        if any("sdg" in s.get("data_sources", []) or "sdg_alignment" in s.get("data_sources", []) for s in template["sections"]):
            data_cache["sdg"] = map_sdg_alignment(company, store)

        if any("five_dimensions" in s.get("data_sources", []) for s in template["sections"]):
            data_cache["five_dim"] = assess_five_dimensions(company, store)

        if any("gap_analysis" in s.get("data_sources", []) for s in template["sections"]):
            data_cache["gaps"] = analyze_gaps(company, store)

        lines = [
            f"{'=' * 60}",
            f"LP DDQ RESPONSE: {template['name']}",
            f"Company: {company.name}",
            f"Sector: {company.sector or 'Not specified'}",
            "Date: Generated by Impact Vision",
            f"{'=' * 60}",
            "",
        ]

        gen_fn = (
            self._generate_narrative_prompt
            if args.narrative_mode == "narrative_prompt"
            else self._generate_section_data
        )

        for section in template["sections"]:
            lines.append(f"--- [{section['id']}] {section['question']} ---")
            lines.append("")
            response = gen_fn(section, company, store, data_cache)
            lines.append(response)
            lines.append("")

        if args.output_format == "json":
            return ToolResult(output=json.dumps({
                "template": template["name"],
                "company": company.name,
                "narrative_mode": args.narrative_mode,
                "sections": [
                    {"id": s["id"], "question": s["question"],
                     "response": gen_fn(s, company, store, data_cache)}
                    for s in template["sections"]
                ],
            }, indent=2))

        if args.output_format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["Template", template["name"]])
            writer.writerow(["Company", company.name])
            writer.writerow([])
            writer.writerow(["Section ID", "Question", "Response"])
            for section in template["sections"]:
                writer.writerow([
                    section["id"],
                    section["question"],
                    self._generate_section_data(section, company, store, data_cache),
                ])
            return ToolResult(output=output.getvalue())

        if args.output_format == "xlsx":
            return self._generate_xlsx(args, template, company, store, data_cache, context)

        text_output = "\n".join(lines)
        if args.draft_review:
            text_output = self._wrap_draft_review(text_output, template["name"], company.name)
        return ToolResult(output=text_output)

    def _generate_xlsx(self, args: LpDdqExportInput, template: dict, company: Company, store, data_cache: dict, context: ToolExecutionContext | None = None) -> ToolResult:
        """Generate an Excel workbook with DDQ responses."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return ToolResult(output="openpyxl required for XLSX export. Install: pip install openpyxl", is_error=True)

        if not args.output_path:
            return ToolResult(output="output_path is required for xlsx format", is_error=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "DDQ Response"

        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        ws.append(["LP DDQ Response", template["name"]])
        ws.append(["Company", company.name])
        ws.append(["Sector", company.sector or "N/A"])
        ws.append(["Date", "Generated by Impact Vision"])
        ws.append([])

        headers = ["Section ID", "Question", "Response", "Data Sources"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for section in template["sections"]:
            response = self._generate_section_data(section, company, store, data_cache)
            ws.append([
                section["id"],
                section["question"],
                response,
                ", ".join(section.get("data_sources", [])),
            ])

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 25

        for row in ws.iter_rows(min_row=7, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        path = Path(args.output_path)
        if not path.is_absolute() and context and hasattr(context, 'cwd'):
            path = context.cwd / path
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

        return ToolResult(
            output=f"XLSX DDQ exported to: {path}\nTemplate: {template['name']}\nCompany: {company.name}\nSections: {len(template['sections'])}",
            metadata={"output_path": str(path), "format": "xlsx"},
        )

    def _generate_section_data(self, section: dict, company: Company, store, data_cache: dict) -> str:
        """Generate data-driven narrative content for a DDQ section."""
        sources = section.get("data_sources", [])
        parts: list[str] = []

        if "impact_thesis" in sources:
            thesis_parts = [f"{company.name}"]
            if company.sector:
                thesis_parts.append(f"operates in the {company.sector} sector")
            if company.geography:
                thesis_parts.append(f"in {company.geography}")
            parts.append(". ".join(thesis_parts) + ".")
            if company.description:
                parts.append(f"\n{company.description[:500]}")
            if company.impact_themes:
                parts.append(f"\nCore impact themes: {', '.join(company.impact_themes)}.")
            if company.sdg_claims:
                parts.append(f"The company contributes to {len(company.sdg_claims)} SDGs: {', '.join(f'SDG {g}' for g in company.sdg_claims)}.")
            if company.stage:
                parts.append(f"Investment stage: {company.stage}.")
            if company.impact_targets:
                parts.append(f"\nImpact targets set for {len(company.impact_targets)} metrics.")
                for t in company.impact_targets:
                    val_str = f"{t.target_value} {t.target_unit}".strip() if t.target_value is not None else t.description
                    parts.append(f"  - {t.metric_id}: {val_str}")

        if "sdg_alignment" in sources and "sdg" in data_cache:
            alignments = data_cache["sdg"]
            top = sorted([a for a in alignments if a.score > 0], key=lambda a: a.score, reverse=True)
            if top:
                parts.append(f"\nSDG Alignment ({len(top)} goals with positive alignment):")
                for a in top[:5]:
                    prov_note = f" [{a.provenance}]" if a.provenance != "evidence-based" else " [evidence-based]"
                    metric_note = f" ({len(a.matched_metrics)} metrics matched)" if a.matched_metrics else ""
                    parts.append(f"  - SDG {a.goal} ({a.goal_name}): {a.score:.0f}/100 [{a.confidence} confidence]{prov_note}{metric_note}")
                if len(top) > 5:
                    parts.append(f"  ... and {len(top) - 5} additional SDG alignments.")

        if "five_dimensions" in sources and "five_dim" in data_cache:
            fd = data_cache["five_dim"]
            parts.append("\n5 Dimensions of Impact Assessment:")
            parts.append(f"  Overall Score: {fd.overall_score:.1f}/5.0 (Grade: {fd.overall_grade})")
            parts.append(f"  Confidence: {fd.overall_provenance}")
            dims = [
                ("What (outcome type)", fd.what),
                ("Who (stakeholders)", fd.who),
                ("How Much (scale/depth)", fd.how_much),
                ("Contribution (additionality)", fd.contribution),
                ("Risk (impact risk)", fd.risk),
            ]
            for label, dim in dims:
                parts.append(f"  - {label}: {dim.score:.1f}/5.0 [{dim.provenance}]")
                if dim.notes:
                    parts.append(f"    {dim.notes[:150]}")

        if "iris_metrics" in sources:
            if company.reported_metrics:
                parts.append(f"\nIRIS+ Metrics Reported ({len(company.reported_metrics)}):")
                for mid, val in list(company.reported_metrics.items())[:10]:
                    m = store.get(mid)
                    name = m.name if m else "Unknown metric"
                    parts.append(f"  - {mid} ({name}): {val}")
                if len(company.reported_metrics) > 10:
                    parts.append(f"  ... and {len(company.reported_metrics) - 10} additional metrics.")
            else:
                parts.append("\nNo IRIS+ metrics currently reported. Recommend establishing baseline measurement.")

        if "gap_analysis" in sources and "gaps" in data_cache:
            gaps = data_cache["gaps"]
            parts.append(f"\nCore Metric Set Coverage: {gaps['coverage_percentage']}%")
            parts.append(f"  Metrics reported: {gaps['metrics_reported']} of {gaps['metrics_reported'] + gaps['metrics_missing']} required")
            if gaps.get("recommendations"):
                parts.append("  Priority actions:")
                for rec in gaps["recommendations"][:3]:
                    parts.append(f"    - {rec}")

        if "dd_checklist" in sources:
            from openharness.impact.dd_checklist import analyze_document_coverage
            if company.description:
                dd_result = analyze_document_coverage(company.description)
                parts.append(f"\nDue Diligence Coverage: {dd_result.coverage_pct}% of {dd_result.total_questions} questions addressed")
                if dd_result.high_priority_gaps:
                    parts.append(f"  High-priority gaps: {len(dd_result.high_priority_gaps)}")
                    for q in dd_result.high_priority_gaps[:3]:
                        parts.append(f"    - {q.question}")

        if "governance" in sources:
            parts.append("\nResponsible Investment Policy:")
            parts.append("  [Insert link or attachment to the firm's RI/ESG policy document]")
            parts.append("  Key elements to address:")
            parts.append("  - ESG integration approach across the investment lifecycle")
            parts.append("  - Exclusion criteria and norms-based screening policies")
            parts.append("  - Voting and engagement policy")
            parts.append("  - Escalation procedures for ESG incidents")

        if "unpri" in sources:
            parts.append("\nFramework Commitments:")
            parts.append("  - UN PRI: [Signatory status / Date of commitment]")
            parts.append("  - GIIN: [Membership status]")
            parts.append("  - Operating Principles for Impact Management: [Alignment status]")
            parts.append("  [Provide PRI Transparency Report reference if applicable]")

        if "measurement_systems" in sources:
            parts.append("\nImpact Measurement Systems:")
            parts.append("  - Measurement framework: IRIS+ / 5 Dimensions of Impact")
            parts.append("  - Data collection: [Describe frequency and methodology]")
            parts.append("  - Third-party verification: [Status of external audit/verification]")
            parts.append("  - Reporting cadence: [Quarterly / Semi-annual / Annual]")

        if "team" in sources:
            parts.append("\nESG Training & Capacity:")
            parts.append("  - [Describe ESG training programs for investment professionals]")
            parts.append("  - [Number of team members with ESG/impact credentials]")
            parts.append("  - [External advisors or committees with ESG expertise]")

        if "outcomes" in sources:
            parts.append("\nImpact Performance & Outcomes:")
            parts.append("  [Insert specific case studies demonstrating impact outcomes]")
            parts.append("  Example format:")
            parts.append("    Portfolio Company: [Name]")
            parts.append("    Impact achieved: [Quantified outcome, e.g., '500 smallholder farmers reached']")
            parts.append("    Measurement method: [How outcome was verified]")
            parts.append("    Attribution: [Fund's contribution to the outcome]")

        if "risk" in sources:
            risk_note = "Impact risk management includes ongoing monitoring of:"
            parts.append(f"\n{risk_note}")
            if company.sector:
                parts.append(f"  - Sector-specific risks for {company.sector}")
            if company.exclusion_flags:
                parts.append(f"  - Exclusion flags: {', '.join(company.exclusion_flags)}")
            else:
                parts.append("  - No exclusion flags identified")
            parts.append("  - [Describe material ESG incidents and responses, if any]")
            parts.append("  - [Describe ESG risk monitoring process during holding period]")

        if any(src in sources for src in ("edci_environment", "edci_social", "edci_governance")):
            parts.append("\nEDCI Reporting:")
            if "edci_environment" in sources:
                env_metrics = {mid: val for mid, val in company.reported_metrics.items() if mid.startswith("OI")}
                parts.append(f"  Environment: {len(env_metrics)} environmental metrics reported")
            if "edci_social" in sources:
                social_metrics = {mid: val for mid, val in company.reported_metrics.items() if mid.startswith("PI")}
                parts.append(f"  Social: {len(social_metrics)} social metrics reported")
            if "edci_governance" in sources:
                parts.append("  Governance: Board composition and oversight data pending")

        if not parts:
            parts.append("[Data not available — provide company description, sector, reported metrics, and SDG claims for comprehensive responses.]")

        return "\n".join(parts)

    def _generate_narrative_prompt(self, section: dict, company: Company, store, data_cache: dict) -> str:
        """Generate a structured prompt for LLM-assisted narrative generation.

        Returns the data context plus writing instructions so the agent can
        produce investor-quality prose for this DDQ section.
        """
        data_content = self._generate_section_data(section, company, store, data_cache)
        question = section.get("question", "")

        prompt_parts = [
            "=== NARRATIVE GENERATION PROMPT ===",
            f"Question: {question}",
            f"Company: {company.name} ({company.sector})",
            "",
            "--- DATA CONTEXT ---",
            data_content,
            "",
            "--- WRITING INSTRUCTIONS ---",
            "Using the data above, write a polished, investor-quality paragraph response to the DDQ question.",
            "Requirements:",
            "- Write in formal third-person prose (not bullet points)",
            "- Reference specific data points (metrics, scores, SDGs) from the context",
            "- Be candid about gaps: acknowledge where data is incomplete",
            "- Use 150-300 words per section",
            "- Maintain factual accuracy: do not invent data not present in the context",
            "- Include quantified outcomes where available",
            "================================",
        ]
        return "\n".join(prompt_parts)

    def _wrap_draft_review(self, output: str, template_name: str, company_name: str) -> str:
        """Wrap output with DRAFT review markers."""
        header = (
            "╔══════════════════════════════════════════════════════════╗\n"
            "║  DRAFT — FOR REVIEW ONLY — NOT FOR DISTRIBUTION        ║\n"
            "╚══════════════════════════════════════════════════════════╝\n"
            f"Template: {template_name}\n"
            f"Company: {company_name}\n"
            "Status: PENDING HUMAN REVIEW\n"
            "Instructions: Review each section for accuracy, completeness,\n"
            "and appropriateness before sharing with LPs.\n"
            "─" * 60 + "\n"
        )
        footer = (
            "\n" + "─" * 60 + "\n"
            "╔══════════════════════════════════════════════════════════╗\n"
            "║  END OF DRAFT — REQUIRES SIGN-OFF BEFORE DISTRIBUTION  ║\n"
            "╚══════════════════════════════════════════════════════════╝"
        )
        return header + output + footer
