"""Tool: Pipeline management for investment deal flow.

CRUD operations for managing companies through pipeline stages,
with CSV/XLSX import/export and funnel analytics.
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field

from openharness.impact.models import PIPELINE_STAGES
from openharness.impact.storage import get_assessment_store
from openharness.tools.base import BaseTool, ToolExecutionContext, ToolResult


class PipelineInput(BaseModel):
    action: Literal[
        "add", "update", "list", "get", "delete",
        "transition", "history", "summary", "dashboard",
        "import_csv", "export_csv", "export_xlsx",
    ] = Field(description="Pipeline operation to perform")
    company_name: str = Field(default="", description="Company name (required for add/update/get/delete/transition/history)")
    pipeline_stage: str = Field(default="sourcing", description=f"Stage: {', '.join(PIPELINE_STAGES)}")
    assigned_to: str = Field(default="", description="Analyst or partner name")
    priority: Literal["high", "medium", "low"] = "medium"
    tags: list[str] = Field(default_factory=list, description="Free-form tags")
    sector: str = ""
    geography: str = ""
    sdg_focus: list[int] = Field(default_factory=list, description="Primary SDG goals")
    investment_size: float | None = Field(default=None, description="Investment amount (USD)")
    notes: str = ""
    rationale: str = Field(default="", description="Rationale for stage transition")
    actor: str = Field(default="", description="Who is making the change")

    filter_stage: str = Field(default="", description="Filter by stage (for list action)")
    filter_sector: str = Field(default="", description="Filter by sector (for list action)")
    filter_sdg: int | None = Field(default=None, description="Filter by SDG goal (for list action)")
    filter_priority: str = Field(default="", description="Filter by priority (for list action)")

    csv_data: str = Field(default="", description="CSV content for import (for import_csv)")
    output_path: str = Field(default="", description="Path for export file")


class PipelineTool(BaseTool):
    name = "pipeline"
    description = (
        "Manage the investment pipeline: add, update, list, filter, and transition "
        "companies through stages (sourcing -> screening -> DD -> IC review -> invested "
        "-> monitoring -> exited/passed). Includes CSV/XLSX import/export."
    )
    input_model = PipelineInput

    def is_read_only(self, arguments: BaseModel) -> bool:
        args = arguments if isinstance(arguments, PipelineInput) else PipelineInput.model_validate(arguments)
        read_only_actions = ("list", "get", "history", "summary")
        if args.action in read_only_actions:
            return True
        if args.action in ("dashboard", "export_csv") and not args.output_path:
            return True
        return False

    async def execute(self, arguments: BaseModel, context: ToolExecutionContext) -> ToolResult:
        args = arguments if isinstance(arguments, PipelineInput) else PipelineInput.model_validate(arguments)
        store = get_assessment_store()

        handlers = {
            "add": self._handle_add,
            "update": self._handle_add,
            "list": self._handle_list,
            "get": self._handle_get,
            "delete": self._handle_delete,
            "transition": self._handle_transition,
            "history": self._handle_history,
            "summary": self._handle_summary,
            "dashboard": self._handle_dashboard,
            "import_csv": self._handle_import_csv,
            "export_csv": self._handle_export_csv,
            "export_xlsx": self._handle_export_xlsx,
        }
        handler = handlers.get(args.action)
        if not handler:
            return ToolResult(output=f"Unknown action: {args.action}", is_error=True)
        return await handler(args, store, context)

    async def _handle_add(self, args: PipelineInput, store, context) -> ToolResult:
        if not args.company_name:
            return ToolResult(output="company_name is required", is_error=True)
        if args.pipeline_stage not in PIPELINE_STAGES:
            return ToolResult(output=f"Invalid stage: {args.pipeline_stage}. Valid: {', '.join(PIPELINE_STAGES)}", is_error=True)
        row_id = store.upsert_pipeline_entry(
            company_name=args.company_name,
            pipeline_stage=args.pipeline_stage,
            assigned_to=args.assigned_to,
            priority=args.priority,
            tags=args.tags,
            sector=args.sector,
            geography=args.geography,
            sdg_focus=args.sdg_focus,
            investment_size=args.investment_size,
            notes=args.notes,
        )
        return ToolResult(
            output=f"Pipeline entry {'updated' if args.action == 'update' else 'created'}: {args.company_name} (stage: {args.pipeline_stage}, id: {row_id})",
            metadata={"id": row_id, "company_name": args.company_name},
        )

    async def _handle_list(self, args: PipelineInput, store, context) -> ToolResult:
        entries = store.list_pipeline(
            stage=args.filter_stage or None,
            sector=args.filter_sector or None,
            sdg=args.filter_sdg,
            priority=args.filter_priority or None,
        )
        if not entries:
            return ToolResult(output="No pipeline entries found matching filters.")
        lines = [f"Pipeline: {len(entries)} companies"]
        for e in entries:
            sdgs = ", ".join(str(s) for s in e.get("sdg_focus", []))
            lines.append(
                f"  [{e['pipeline_stage'].upper():15s}] {e['company_name']}"
                f" | {e.get('sector', '')} | {e.get('geography', '')}"
                f" | SDGs: {sdgs or 'N/A'} | Priority: {e.get('priority', 'medium')}"
            )
        return ToolResult(output="\n".join(lines), metadata={"count": len(entries)})

    async def _handle_get(self, args: PipelineInput, store, context) -> ToolResult:
        if not args.company_name:
            return ToolResult(output="company_name is required", is_error=True)
        entry = store.get_pipeline_entry(args.company_name)
        if not entry:
            return ToolResult(output=f"Not found: {args.company_name}")
        transitions = store.get_transitions(args.company_name)
        lines = [
            f"Company: {entry['company_name']}",
            f"Stage: {entry['pipeline_stage']}",
            f"Assigned: {entry.get('assigned_to', 'N/A')}",
            f"Priority: {entry.get('priority', 'medium')}",
            f"Sector: {entry.get('sector', 'N/A')}",
            f"Geography: {entry.get('geography', 'N/A')}",
            f"SDG Focus: {', '.join(str(s) for s in entry.get('sdg_focus', []))}",
            f"Investment: ${entry['investment_size']:,.0f}" if entry.get('investment_size') else "Investment: N/A",
            f"Tags: {', '.join(entry.get('tags', []))}",
            f"Notes: {entry.get('notes', '')}",
        ]
        if transitions:
            lines.append(f"\nTransition History ({len(transitions)}):")
            for t in transitions:
                lines.append(f"  {t['timestamp'][:10]}: {t['from_stage'] or '(new)'} -> {t['to_stage']}"
                             f" | {t['actor'] or 'system'}: {t['rationale']}")
        return ToolResult(output="\n".join(lines), metadata={"entry": entry})

    async def _handle_delete(self, args: PipelineInput, store, context) -> ToolResult:
        if not args.company_name:
            return ToolResult(output="company_name is required", is_error=True)
        ok = store.delete_pipeline_entry(args.company_name)
        msg = f"Deleted: {args.company_name}" if ok else f"Not found: {args.company_name}"
        return ToolResult(output=msg)

    async def _handle_transition(self, args: PipelineInput, store, context) -> ToolResult:
        if not args.company_name:
            return ToolResult(output="company_name is required", is_error=True)
        if args.pipeline_stage not in PIPELINE_STAGES:
            return ToolResult(output=f"Invalid stage: {args.pipeline_stage}", is_error=True)
        ok = store.transition_stage(
            company_name=args.company_name,
            new_stage=args.pipeline_stage,
            actor=args.actor,
            rationale=args.rationale,
            notes=args.notes,
        )
        if not ok:
            return ToolResult(output=f"Not found in pipeline: {args.company_name}", is_error=True)
        return ToolResult(
            output=f"Transitioned {args.company_name} to {args.pipeline_stage} (by {args.actor or 'system'})",
        )

    async def _handle_history(self, args: PipelineInput, store, context) -> ToolResult:
        if not args.company_name:
            return ToolResult(output="company_name is required", is_error=True)
        transitions = store.get_transitions(args.company_name)
        if not transitions:
            return ToolResult(output=f"No transitions found for {args.company_name}")
        lines = [f"Stage transitions for {args.company_name}: ({len(transitions)})"]
        for t in transitions:
            lines.append(
                f"  {t['timestamp'][:10]}: {t['from_stage'] or '(new)'} -> {t['to_stage']}"
                f" | {t['actor'] or 'system'}: {t['rationale']}"
            )
        return ToolResult(output="\n".join(lines))

    async def _handle_summary(self, args: PipelineInput, store, context) -> ToolResult:
        summary = store.pipeline_summary()
        if not summary:
            return ToolResult(output="Pipeline is empty")
        total = sum(summary.values())
        lines = [f"Pipeline Summary ({total} companies):"]
        for stage in PIPELINE_STAGES:
            count = summary.get(stage, 0)
            bar = "█" * count + "░" * (max(0, 10 - count))
            lines.append(f"  {stage:17s} {bar} {count}")
        return ToolResult(output="\n".join(lines), metadata={"summary": summary, "total": total})

    async def _handle_dashboard(self, args: PipelineInput, store, context) -> ToolResult:
        """Generate an HTML pipeline dashboard with funnel, sectors, and SDG coverage."""
        entries = store.list_pipeline()
        summary = store.pipeline_summary()
        if not entries:
            return ToolResult(output="Pipeline is empty. Add companies first.")

        total = len(entries)
        sectors: dict[str, int] = {}
        sdgs: dict[int, int] = {}
        priorities: dict[str, int] = {}
        geos: dict[str, int] = {}
        for e in entries:
            s = e.get("sector", "") or "Unknown"
            sectors[s] = sectors.get(s, 0) + 1
            for sdg in e.get("sdg_focus", []):
                sdgs[sdg] = sdgs.get(sdg, 0) + 1
            p = e.get("priority", "medium")
            priorities[p] = priorities.get(p, 0) + 1
            g = e.get("geography", "") or "Unknown"
            geos[g] = geos.get(g, 0) + 1

        html_parts = [
            '<!DOCTYPE html><html><head><meta charset="UTF-8">',
            '<title>Pipeline Dashboard</title>',
            '<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>',
            '<style>',
            'body{font-family:Inter,-apple-system,sans-serif;max-width:1080px;margin:0 auto;padding:24px;background:#f5f7fa;color:#1a1a2e}',
            'h1{color:#0d47a1}h2{color:#0d47a1;border-bottom:2px solid #e3f2fd;padding-bottom:8px}',
            '.cards{display:flex;gap:16px;flex-wrap:wrap;margin:16px 0}',
            '.card{background:white;border-radius:12px;padding:20px 28px;box-shadow:0 1px 3px rgba(0,0,0,0.08);text-align:center;min-width:120px;border:1px solid #e0e4e8}',
            '.card .val{font-size:2em;font-weight:700;color:#0d47a1}.card .lbl{font-size:0.8em;color:#5f6368;text-transform:uppercase}',
            '.chart-row{display:flex;gap:20px;flex-wrap:wrap;margin:16px 0}',
            '.chart-box{flex:1 1 420px;background:white;border-radius:12px;padding:20px;box-shadow:0 1px 3px rgba(0,0,0,0.08);border:1px solid #e0e4e8}',
            'table{border-collapse:collapse;width:100%;margin:12px 0;background:white;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,0.08)}',
            'th{background:#0d47a1;color:white;padding:10px 14px;text-align:left;font-size:0.8em;text-transform:uppercase}',
            'td{padding:8px 14px;border-bottom:1px solid #e0e4e8;font-size:0.88em}',
            '.stage-badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:0.75em;font-weight:600;text-transform:uppercase}',
            '</style></head><body>',
            '<h1>Pipeline Dashboard</h1>',
            f'<p style="color:#5f6368">{total} companies in pipeline</p>',
        ]

        html_parts.append('<div class="cards">')
        stage_colors = {
            "sourcing": "#1976d2", "screening": "#7b1fa2", "dd_in_progress": "#f57c00",
            "ic_review": "#c62828", "invested": "#2e7d32", "monitoring": "#00695c",
            "exited": "#455a64", "passed": "#9e9e9e",
        }
        for stage in PIPELINE_STAGES:
            cnt = summary.get(stage, 0)
            color = stage_colors.get(stage, "#666")
            label = stage.replace("_", " ").title()
            html_parts.append(f'<div class="card"><div class="val" style="color:{color}">{cnt}</div><div class="lbl">{label}</div></div>')
        html_parts.append('</div>')

        # Funnel chart
        funnel_stages = [s.replace("_", " ").title() for s in PIPELINE_STAGES]
        funnel_vals = [str(summary.get(s, 0)) for s in PIPELINE_STAGES]
        html_parts.append('<div class="chart-row"><div class="chart-box" id="funnel-chart"></div>')
        html_parts.append('<div class="chart-box" id="sector-chart"></div></div>')
        html_parts.append(f'''<script>
Plotly.newPlot('funnel-chart', [{{type:'funnel',y:{json.dumps(funnel_stages)},x:[{",".join(funnel_vals)}],
marker:{{color:{json.dumps([stage_colors.get(s,"#666") for s in PIPELINE_STAGES])}}}
}}],{{title:'Pipeline Funnel',height:350,margin:{{l:120,r:20,t:40,b:20}},font:{{family:'Inter,sans-serif'}}}},{{responsive:true}});
Plotly.newPlot('sector-chart',[{{type:'pie',labels:{json.dumps(list(sectors.keys()))},values:{json.dumps(list(sectors.values()))},
hole:0.4}}],{{title:'By Sector',height:350,margin:{{l:20,r:20,t:40,b:20}},font:{{family:'Inter,sans-serif'}}}},{{responsive:true}});
</script>''')

        if sdgs:
            sdg_labels = [f"SDG {g}" for g in sorted(sdgs)]
            sdg_vals = [str(sdgs[g]) for g in sorted(sdgs)]
            html_parts.append('<div class="chart-row"><div class="chart-box" id="sdg-pipe-chart"></div></div>')
            html_parts.append(f'''<script>
Plotly.newPlot('sdg-pipe-chart',[{{type:'bar',x:{json.dumps(sdg_labels)},y:[{",".join(sdg_vals)}],
marker:{{color:'#1976d2',cornerradius:4}}}}],
{{title:'SDG Focus Distribution',height:300,margin:{{l:40,r:20,t:40,b:40}},
font:{{family:'Inter,sans-serif'}}}},{{responsive:true}});
</script>''')

        html_parts.append('<h2>All Companies</h2>')
        html_parts.append('<table><tr><th>Company</th><th>Stage</th><th>Sector</th><th>Geography</th><th>SDGs</th><th>Priority</th></tr>')
        for e in entries:
            stage = e["pipeline_stage"]
            color = stage_colors.get(stage, "#666")
            sdg_str = ", ".join(str(s) for s in e.get("sdg_focus", []))
            html_parts.append(
                f'<tr><td><strong>{e["company_name"]}</strong></td>'
                f'<td><span class="stage-badge" style="background:{color}20;color:{color}">{stage.replace("_"," ").title()}</span></td>'
                f'<td>{e.get("sector","")}</td><td>{e.get("geography","")}</td>'
                f'<td>{sdg_str}</td><td>{e.get("priority","medium")}</td></tr>'
            )
        html_parts.append('</table></body></html>')

        html = "\n".join(html_parts)
        if args.output_path:
            path = Path(args.output_path)
            if not path.is_absolute():
                path = context.cwd / path
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(html, encoding="utf-8")
            return ToolResult(output=f"Dashboard saved to {path} ({total} companies)")
        return ToolResult(
            output=f"Pipeline dashboard generated ({total} companies). Provide output_path to save HTML.",
            metadata={"html_length": len(html), "total": total},
        )

    async def _handle_import_csv(self, args: PipelineInput, store, context) -> ToolResult:
        csv_content = args.csv_data
        if not csv_content and args.output_path:
            path = Path(args.output_path)
            if not path.is_absolute():
                path = context.cwd / path
            if path.exists():
                csv_content = path.read_text(encoding="utf-8")
        if not csv_content:
            return ToolResult(output="Provide csv_data or output_path with CSV content", is_error=True)
        reader = csv.DictReader(io.StringIO(csv_content))
        imported = 0
        for row in reader:
            name = row.get("company_name", "").strip()
            if not name:
                continue
            sdgs = []
            raw_sdg = row.get("sdg_focus", "")
            if raw_sdg:
                sdgs = [int(s.strip()) for s in raw_sdg.split(",") if s.strip().isdigit()]
            inv = None
            raw_inv = row.get("investment_size", "")
            if raw_inv:
                try:
                    inv = float(raw_inv.replace(",", ""))
                except ValueError:
                    pass
            store.upsert_pipeline_entry(
                company_name=name,
                pipeline_stage=row.get("pipeline_stage", "sourcing"),
                assigned_to=row.get("assigned_to", ""),
                priority=row.get("priority", "medium"),
                tags=[t.strip() for t in row.get("tags", "").split(",") if t.strip()],
                sector=row.get("sector", ""),
                geography=row.get("geography", ""),
                sdg_focus=sdgs,
                investment_size=inv,
                notes=row.get("notes", ""),
            )
            imported += 1
        return ToolResult(output=f"Imported {imported} pipeline entries from CSV")

    async def _handle_export_csv(self, args: PipelineInput, store, context) -> ToolResult:
        entries = store.list_pipeline()
        if not entries:
            return ToolResult(output="Pipeline is empty, nothing to export")
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=[
            "company_name", "pipeline_stage", "assigned_to", "priority",
            "sector", "geography", "sdg_focus", "investment_size", "tags", "notes",
        ])
        writer.writeheader()
        for e in entries:
            writer.writerow({
                "company_name": e["company_name"],
                "pipeline_stage": e["pipeline_stage"],
                "assigned_to": e.get("assigned_to", ""),
                "priority": e.get("priority", ""),
                "sector": e.get("sector", ""),
                "geography": e.get("geography", ""),
                "sdg_focus": ",".join(str(s) for s in e.get("sdg_focus", [])),
                "investment_size": e.get("investment_size", ""),
                "tags": ",".join(e.get("tags", [])),
                "notes": e.get("notes", ""),
            })
        output = buf.getvalue()
        if args.output_path:
            path = Path(args.output_path)
            if not path.is_absolute():
                path = context.cwd / path
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(output, encoding="utf-8")
            return ToolResult(output=f"Exported {len(entries)} entries to {path}")
        return ToolResult(output=output, metadata={"count": len(entries)})

    async def _handle_export_xlsx(self, args: PipelineInput, store, context) -> ToolResult:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return ToolResult(output="openpyxl required. Install: pip install openpyxl", is_error=True)

        entries = store.list_pipeline()
        if not entries:
            return ToolResult(output="Pipeline is empty")

        path = Path(args.output_path) if args.output_path else Path("pipeline_export.xlsx")
        if not path.is_absolute():
            path = context.cwd / path

        wb = Workbook()
        ws = wb.active
        ws.title = "Pipeline"
        headers = ["Company", "Stage", "Assigned To", "Priority", "Sector",
                    "Geography", "SDG Focus", "Investment Size", "Tags", "Notes"]
        ws.append(headers)
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for e in entries:
            ws.append([
                e["company_name"], e["pipeline_stage"], e.get("assigned_to", ""),
                e.get("priority", ""), e.get("sector", ""), e.get("geography", ""),
                ", ".join(str(s) for s in e.get("sdg_focus", [])),
                e.get("investment_size", ""), ", ".join(e.get("tags", [])),
                e.get("notes", ""),
            ])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return ToolResult(output=f"Exported {len(entries)} entries to {path}")
