"""Tool: Generate impact assessment reports (HTML, CSV, JSON)."""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal

from pydantic import BaseModel, Field

from openharness.impact.database import get_metric_store
from openharness.impact.five_dimensions import assess_five_dimensions
from openharness.impact.gap_analysis import analyze_gaps
from openharness.impact.greenwashing import assess_greenwashing
from openharness.impact.models import Company
from openharness.impact.sdg_mapper import generate_sdg_gap_recommendations, map_sdg_alignment
from openharness.tools.base import BaseTool, ToolExecutionContext, ToolResult

_SECTOR_OPPORTUNITIES: dict[str, list[str]] = {
    "agriculture": [
        "Food security: improving access to nutritious food for underserved populations",
        "Poverty reduction: increasing income for smallholder farmers",
        "Sustainable supply chains: promoting responsible sourcing and production",
        "Climate adaptation: developing climate-resilient farming practices",
        "Rural employment: creating jobs in farming communities",
    ],
    "livestock": [
        "Food security: providing protein and nutrition to local markets",
        "Poverty reduction: supporting livelihoods of small-scale farmers",
        "Economic development: strengthening local agricultural value chains",
        "Sustainable practices: implementing responsible animal husbandry",
        "Rural employment: creating jobs in farming and processing",
    ],
    "healthcare": [
        "Health outcomes: improving access to quality healthcare services",
        "Health equity: reducing disparities in healthcare access",
        "Disease prevention: supporting public health initiatives",
        "Workforce development: training healthcare professionals",
    ],
    "energy": [
        "Clean energy access: providing affordable clean energy to underserved areas",
        "Climate mitigation: reducing greenhouse gas emissions",
        "Energy independence: reducing reliance on fossil fuels",
        "Job creation: creating green energy employment opportunities",
    ],
    "education": [
        "Educational access: reaching underserved or marginalized populations",
        "Skills development: building employable skills for the workforce",
        "Digital inclusion: bridging the digital divide",
        "Gender equity: improving educational access for girls and women",
    ],
    "fintech": [
        "Financial inclusion: providing access to financial services for the unbanked",
        "Economic empowerment: enabling savings, credit, and insurance",
        "Gender equity: improving financial access for women",
        "Efficiency: reducing transaction costs for low-income users",
    ],
    "water": [
        "Clean water access: providing safe drinking water to underserved communities",
        "Sanitation: improving hygiene and reducing waterborne disease",
        "Water efficiency: promoting sustainable water management practices",
    ],
    "technology": [
        "Digital inclusion: bridging the digital divide for underserved populations",
        "Efficiency gains: improving productivity and reducing waste",
        "Innovation: enabling new solutions to social and environmental challenges",
    ],
}

_SECTOR_RISKS: dict[str, list[str]] = {
    "agriculture": [
        "Environmental degradation: soil depletion, water pollution from fertilizers/pesticides",
        "Climate vulnerability: sensitivity to extreme weather events and climate change",
        "Labor conditions: risk of exploitative labor practices",
        "Land use change: deforestation or biodiversity loss from land conversion",
        "Market volatility: commodity price fluctuations affecting farmer incomes",
    ],
    "livestock": [
        "Environmental pollution: waste management, methane emissions, water contamination",
        "Animal welfare: risk of poor animal husbandry practices",
        "Disease risk: zoonotic disease transmission and antibiotic resistance",
        "Carbon footprint: significant greenhouse gas emissions from livestock",
        "Resource intensity: high water and feed consumption per unit of protein",
        "Community impact: odor, waste runoff affecting neighboring communities",
    ],
    "healthcare": [
        "Access inequality: risk of services remaining unaffordable for the poorest",
        "Quality variance: inconsistent quality of care across locations",
        "Data privacy: risks around patient health data security",
    ],
    "energy": [
        "Environmental impact: land use, waste from equipment, resource extraction",
        "Community displacement: risk of displacing communities for energy projects",
        "Technology risk: rapidly changing technology making investments obsolete",
    ],
    "education": [
        "Quality risk: providing access without ensuring quality outcomes",
        "Digital divide: technology-dependent models excluding the most vulnerable",
        "Sustainability: dependence on grants or subsidies for viability",
    ],
    "fintech": [
        "Over-indebtedness: risk of predatory lending to vulnerable populations",
        "Data privacy: risks around financial data security and misuse",
        "Digital exclusion: services inaccessible to those without smartphones/internet",
    ],
    "water": [
        "Sustainability: risk of depleting water sources without replenishment",
        "Infrastructure maintenance: long-term maintenance of water systems",
        "Affordability: pricing that excludes the poorest communities",
    ],
    "technology": [
        "Digital divide: reinforcing existing inequalities through technology access",
        "Privacy and surveillance: risks around data collection and misuse",
        "Job displacement: automation reducing employment opportunities",
    ],
}


def _infer_opportunities_and_risks(company: Company) -> dict[str, list[str]]:
    """Infer impact opportunities and risks from sector and description."""
    text = f"{company.description} {company.sector}".lower()
    opportunities: list[str] = []
    risks: list[str] = []

    for sector_key in _SECTOR_OPPORTUNITIES:
        if sector_key in text:
            opportunities.extend(_SECTOR_OPPORTUNITIES[sector_key])
    for sector_key in _SECTOR_RISKS:
        if sector_key in text:
            risks.extend(_SECTOR_RISKS[sector_key])

    if not opportunities:
        opportunities = ["Further analysis needed to identify specific impact opportunities"]
    if not risks:
        risks = ["Further analysis needed to identify specific impact risks"]

    return {"opportunities": list(dict.fromkeys(opportunities)), "risks": list(dict.fromkeys(risks))}


class ImpactReportInput(BaseModel):
    company_name: str = Field(description="Name of the company")
    company_description: str = Field(default="")
    sector: str = Field(default="")
    geography: str = Field(default="", description="Country or region")
    impact_themes: list[str] = Field(default_factory=list)
    reported_metrics: dict[str, str] = Field(default_factory=dict)
    sdg_claims: list[int] = Field(default_factory=list)
    output_format: Literal["html", "csv", "json", "text", "xlsx", "pdf"] = Field(
        default="text",
        description="Output format for the report ('xlsx' for Excel, 'pdf' for print-ready)"
    )
    output_path: str = Field(
        default="",
        description="File path to save the report (optional; if empty, returns as text)",
    )
    include_gap_analysis: bool = Field(default=True)
    include_sdg_mapping: bool = Field(default=True)
    include_five_dimensions: bool = Field(default=True)
    report_type: Literal["full", "target_progress", "lp_ready"] = Field(
        default="full",
        description=(
            "'full': Complete assessment report. "
            "'target_progress': Focused report on target progress and trajectory projections. "
            "'lp_ready': LP-formatted individual company report with executive summary."
        ),
    )
    narrative_mode: Literal["data", "narrative_prompt"] = Field(
        default="data",
        description=(
            "'data': Standard data-driven output. "
            "'narrative_prompt': Append LLM writing prompts for executive summary, "
            "key findings, and recommendations so the agent can generate polished narratives."
        ),
    )
    draft_review: bool = Field(
        default=False,
        description="If True, output is wrapped with DRAFT markers for human review.",
    )
    compare_assessment_id: str = Field(
        default="",
        description="Optional previous assessment ID for side-by-side comparison.",
    )


class ImpactReportTool(BaseTool):
    name = "impact_report"
    description = (
        "Generate a comprehensive impact assessment report for a company. "
        "Includes 5-Dimension scoring, SDG alignment mapping, and gap analysis. "
        "Supports HTML, CSV, JSON, and text output formats. "
        "Optionally saves to a file."
    )
    input_model = ImpactReportInput

    def is_read_only(self, arguments: BaseModel) -> bool:
        return True

    async def execute(self, arguments: BaseModel, context: ToolExecutionContext) -> ToolResult:
        args = arguments if isinstance(arguments, ImpactReportInput) else ImpactReportInput.model_validate(arguments)

        try:
            store = get_metric_store()
        except FileNotFoundError as e:
            return ToolResult(output=str(e), is_error=True)

        company = Company(
            name=args.company_name,
            description=args.company_description,
            sector=args.sector,
            geography=args.geography,
            impact_themes=args.impact_themes,
            reported_metrics=args.reported_metrics,
            sdg_claims=args.sdg_claims,
        )

        report_data: dict = {
            "company": company.model_dump(),
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "catalog_version": "IRIS+ 5.3c",
        }

        if args.include_five_dimensions:
            fd_result = assess_five_dimensions(company, store)
            report_data["five_dimensions"] = fd_result.model_dump()

        if args.include_sdg_mapping:
            sdg_results = map_sdg_alignment(company, store)
            sdg_recs = generate_sdg_gap_recommendations(sdg_results, company, store)
            sdg_dicts = []
            for a in sdg_results:
                d = a.model_dump()
                d["recommendations"] = sdg_recs.get(a.goal, [])
                sdg_dicts.append(d)
            report_data["sdg_alignments"] = sdg_dicts

        if args.include_gap_analysis:
            gap_result = analyze_gaps(company, store)
            report_data["gap_analysis"] = gap_result

        report_data["impact_analysis"] = _infer_opportunities_and_risks(company)
        gw = assess_greenwashing(company)
        gw_dump = gw.model_dump()
        # Build sub_scores dict for report rendering compatibility
        gw_dump["sub_scores"] = {
            "claim_metric_gap": gw.claim_metric_gap,
            "adverse_omission": gw.adverse_omission,
            "specificity": gw.specificity,
            "selectivity": gw.selectivity,
            "verification": gw.verification,
        }
        report_data["greenwashing"] = gw_dump

        if company.impact_targets:
            from openharness.impact.trend_analysis import assess_target_progress
            report_data["target_tracking"] = assess_target_progress(company)

        if company.beneficiary_feedback:
            report_data["beneficiary_feedback"] = company.beneficiary_feedback.model_dump()

        from openharness.impact.benchmarks import compare_to_benchmark
        if "five_dimensions" in report_data and company.sector:
            fd = report_data["five_dimensions"]
            five_d_scores = {
                "what": fd["what"]["score"],
                "who": fd["who"]["score"],
                "how_much": fd["how_much"]["score"],
                "contribution": fd["contribution"]["score"],
                "risk": fd["risk"]["score"],
            }
            coverage = report_data.get("gap_analysis", {}).get("coverage_percentage", 0)
            bm = compare_to_benchmark(company.sector, five_d_scores, fd["overall_score"], coverage)
            if bm.get("benchmark_available"):
                report_data["benchmark_comparison"] = bm

        if args.compare_assessment_id:
            report_data["comparison"] = _load_comparison_data(
                args.compare_assessment_id, report_data
            )

        if args.output_format == "xlsx":
            return _to_xlsx(report_data, args.output_path, context)
        elif args.output_format == "json":
            output = json.dumps(report_data, indent=2, default=str)
        elif args.output_format == "csv":
            output = _to_csv(report_data)
        elif args.output_format in ("html", "pdf"):
            output = _to_html(report_data)
        elif args.report_type == "target_progress":
            output = _to_target_progress_text(report_data)
        elif args.report_type == "lp_ready":
            output = _to_lp_ready_text(report_data)
        else:
            output = _to_text(report_data)

        if args.narrative_mode == "narrative_prompt":
            output += "\n\n" + _generate_report_narrative_prompt(report_data)

        if args.draft_review:
            output = _wrap_report_draft(output, company.name)

        if args.output_format == "pdf":
            return _to_pdf(output, args.output_path, context)

        if args.output_path:
            path = Path(args.output_path)
            if not path.is_absolute():
                path = context.cwd / path
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(output, encoding="utf-8")
            summary = _to_text(report_data) if args.output_format != "text" else output[:1500]
            return ToolResult(
                output=f"Report saved to: {path}\nFormat: {args.output_format}\n\n{summary}",
                metadata={"output_path": str(path), "format": args.output_format},
            )

        if args.output_format == "html" and len(output) > 2000:
            summary = _to_text(report_data)
            return ToolResult(
                output=f"HTML report generated ({len(output)} chars). Text summary:\n\n{summary}",
                metadata={"format": args.output_format, "html_length": len(output)},
            )

        return ToolResult(
            output=output,
            metadata={"format": args.output_format},
        )


def _to_text(data: dict) -> str:
    lines = [
        "=" * 70,
        f"IMPACT ASSESSMENT REPORT: {data['company']['name']}",
        f"Generated: {data['generated_at']}",
        f"Standard: {data['catalog_version']}",
        "=" * 70,
        "",
    ]

    if "five_dimensions" in data:
        fd = data["five_dimensions"]
        lines.append("5 DIMENSIONS OF IMPACT")
        lines.append("-" * 40)
        lines.append(f"Overall Grade: {fd['overall_grade']} ({fd['overall_score']}/5.0)")
        for dim_name in ["what", "who", "how_much", "contribution", "risk"]:
            dim = fd[dim_name]
            lines.append(f"  {dim['dimension']}: {dim['score']}/5.0 | {dim['notes']}")
        if fd.get("recommendations"):
            lines.append("\nRecommendations:")
            for r in fd["recommendations"]:
                lines.append(f"  - {r}")
        lines.append("")

    if "sdg_alignments" in data:
        lines.append("SDG ALIGNMENT")
        lines.append("-" * 40)
        for a in data["sdg_alignments"]:
            if a["score"] > 0:
                lines.append(f"  SDG {a['goal']} ({a['goal_name']}): {a['score']}/100 [{a['confidence']}]")
        lines.append("")

    if "gap_analysis" in data:
        ga = data["gap_analysis"]
        lines.append("GAP ANALYSIS")
        lines.append("-" * 40)
        lines.append(f"  Coverage: {ga['coverage_percentage']}% ({ga['metrics_reported']}/{ga['core_metric_set_size']})")
        if ga.get("missing"):
            lines.append("  Missing:")
            for m in ga["missing"][:10]:
                lines.append(f"    - {m['id']}: {m['name']}")
        lines.append("")

    if "impact_analysis" in data:
        ia = data["impact_analysis"]
        lines.append("IMPACT OPPORTUNITIES")
        lines.append("-" * 40)
        for o in ia.get("opportunities", []):
            lines.append(f"  + {o}")
        lines.append("")
        lines.append("IMPACT RISKS")
        lines.append("-" * 40)
        for r in ia.get("risks", []):
            lines.append(f"  ! {r}")
        lines.append("")

    if "greenwashing" in data:
        gw = data["greenwashing"]
        lines.append("GREENWASHING / IMPACT-WASHING RISK")
        lines.append("-" * 40)
        lines.append(f"  Risk Score: {gw.get('overall_score', 0)}/100 — {gw.get('classification', 'Unknown')}")
        for sname, sval in gw.get("sub_scores", {}).items():
            lines.append(f"    {sname.replace('_', ' ').title()}: {sval}/100")
        for flag in gw.get("flags", []):
            lines.append(f"  ! {flag}")
        for rec in gw.get("recommendations", []):
            lines.append(f"  > {rec}")
        lines.append("")

    if "benchmark_comparison" in data:
        bm = data["benchmark_comparison"]
        lines.append("SECTOR BENCHMARK COMPARISON")
        lines.append("-" * 40)
        lines.append(f"  Sector: {bm['sector']} ({bm['sample_note']})")
        ov = bm["overall"]
        arrow = "^" if ov["delta"] > 0 else ("v" if ov["delta"] < 0 else "=")
        lines.append(f"  Overall: {ov['actual']:.1f} vs {ov['benchmark']:.1f} benchmark ({arrow} {ov['delta']:+.1f})")
        for dim, vals in bm["dimensions"].items():
            arrow = "^" if vals["delta"] > 0 else ("v" if vals["delta"] < 0 else "=")
            lines.append(f"    {dim}: {vals['actual']:.1f} vs {vals['benchmark']:.1f} ({arrow} {vals['delta']:+.1f})")
        cov = bm["coverage"]
        lines.append(f"  Coverage: {cov['actual']:.0f}% vs {cov['benchmark']:.0f}% benchmark ({cov['delta']:+.1f}%)")
        lines.append("")

    return "\n".join(lines)


def _generate_report_narrative_prompt(data: dict) -> str:
    """Generate structured LLM prompts for executive summary, findings, and recommendations."""
    company = data.get("company", {})
    fd = data.get("five_dimensions", {})
    sdg = data.get("sdg_alignments", [])
    gaps = data.get("gap_analysis", {})
    gw = data.get("greenwashing", {})

    sections = []

    sections.append("=== NARRATIVE GENERATION PROMPTS ===\n")
    sections.append("The following prompts provide structured context for the agent to generate")
    sections.append("investor-quality narratives. Each section includes data points and writing instructions.\n")

    # Executive Summary prompt
    exec_data = [f"Company: {company.get('name', 'Unknown')} ({company.get('sector', 'N/A')})"]
    if company.get("geography"):
        exec_data.append(f"Geography: {company['geography']}")
    if fd:
        exec_data.append(f"5D Score: {fd.get('overall_score', 'N/A')}/5 (Grade: {fd.get('overall_grade', 'N/A')})")
    top_sdgs = sorted(sdg, key=lambda s: s.get("score", 0), reverse=True)[:3]
    if top_sdgs:
        sdg_str = ", ".join(f"SDG {s['goal']} ({s.get('score', 0):.0f}/100)" for s in top_sdgs)
        exec_data.append(f"Top SDGs: {sdg_str}")
    if gaps:
        exec_data.append(f"Core Metric Coverage: {gaps.get('coverage_percentage', 0)}%")
    if isinstance(gw, dict):
        exec_data.append(f"Greenwashing Risk: {gw.get('overall_score', 'N/A')}/100 ({gw.get('classification', 'N/A')})")

    sections.append("--- EXECUTIVE SUMMARY ---")
    sections.append("Data points:")
    for d_item in exec_data:
        sections.append(f"  - {d_item}")
    sections.append("Instructions: Write a 200-word executive summary covering the company's impact profile,")
    sections.append("key strengths, primary risks, and recommended next steps. Use formal investor language.\n")

    # Key Findings prompt
    findings_data = []
    if fd:
        dims = fd.get("dimension_averages", {}) if "dimension_averages" in fd else {}
        if not dims:
            for dim_name in ("what", "who", "how_much", "contribution", "risk"):
                dim = fd.get(dim_name, {})
                if isinstance(dim, dict) and "score" in dim:
                    findings_data.append(f"{dim_name}: {dim['score']}/5 [{dim.get('provenance', 'N/A')}]")
    if gaps:
        findings_data.append(f"Metrics reported: {gaps.get('metrics_reported', 0)}/{gaps.get('metrics_reported', 0) + gaps.get('metrics_missing', 0)}")
        recs = gaps.get("recommendations", [])[:3]
        for r in recs:
            findings_data.append(f"Gap recommendation: {r}")

    sections.append("--- KEY FINDINGS ---")
    sections.append("Data points:")
    for f_item in findings_data:
        sections.append(f"  - {f_item}")
    sections.append("Instructions: Summarize 3-5 key findings about the company's impact measurement maturity,")
    sections.append("data quality, and alignment with international standards. Be specific and data-driven.\n")

    # Recommendations prompt
    sections.append("--- RECOMMENDATIONS ---")
    sections.append("Instructions: Based on the full report data above, write 3-5 prioritized recommendations")
    sections.append("for improving the company's impact measurement and reporting. Each recommendation should:")
    sections.append("  1. Identify the specific gap or weakness")
    sections.append("  2. Propose a concrete action")
    sections.append("  3. Reference the relevant framework or standard (IRIS+, SDG, SASB, etc.)")
    sections.append("  4. Estimate the effort level (quick win / medium-term / strategic)")
    sections.append("================================\n")

    return "\n".join(sections)


def _wrap_report_draft(output: str, company_name: str) -> str:
    """Wrap report output with DRAFT review markers."""
    header = (
        "╔══════════════════════════════════════════════════════════════════╗\n"
        "║  DRAFT IMPACT REPORT — FOR INTERNAL REVIEW ONLY               ║\n"
        "╚══════════════════════════════════════════════════════════════════╝\n"
        f"Company: {company_name}\n"
        "Status: PENDING HUMAN REVIEW\n"
        "Instructions: Review all scores, claims, and recommendations for\n"
        "accuracy before sharing with stakeholders or including in LP reports.\n"
        "─" * 66 + "\n"
    )
    footer = (
        "\n" + "─" * 66 + "\n"
        "╔══════════════════════════════════════════════════════════════════╗\n"
        "║  END OF DRAFT — REQUIRES INVESTMENT TEAM SIGN-OFF             ║\n"
        "╚══════════════════════════════════════════════════════════════════╝"
    )
    return header + output + footer


def _to_csv(data: dict) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(["Section", "Metric", "Value", "Details"])
    writer.writerow(["Company", "Name", data["company"]["name"], ""])
    writer.writerow(["Company", "Generated", data["generated_at"], ""])

    if "five_dimensions" in data:
        fd = data["five_dimensions"]
        writer.writerow(["5D", "Overall Grade", fd["overall_grade"], f"{fd['overall_score']}/5.0"])
        for dim_name in ["what", "who", "how_much", "contribution", "risk"]:
            dim = fd[dim_name]
            writer.writerow(["5D", dim["dimension"], f"{dim['score']}/5.0", dim["notes"]])

    if "sdg_alignments" in data:
        for a in data["sdg_alignments"]:
            if a["score"] > 0:
                writer.writerow([
                    "SDG", f"SDG {a['goal']}", f"{a['score']}/100",
                    f"{a['confidence']} | metrics: {','.join(a.get('matched_metrics', [])[:3])}"
                ])

    if "gap_analysis" in data:
        ga = data["gap_analysis"]
        writer.writerow(["Gap", "Coverage", f"{ga['coverage_percentage']}%", ""])
        for m in ga.get("missing", []):
            writer.writerow(["Gap", m["id"], "MISSING", m["name"]])
        for m in ga.get("reported", []):
            writer.writerow(["Gap", m["id"], m.get("value", ""), m["name"]])

    return buf.getvalue()


def _interactive_scoring_section(fd: dict, sdg_alignments: list, company_name: str = "") -> str:
    """Generate an interactive HTML section with checkboxes that adjust scores."""
    base_scores = {
        "what": fd["what"]["score"],
        "who": fd["who"]["score"],
        "how_much": fd["how_much"]["score"],
        "contribution": fd["contribution"]["score"],
        "risk": fd["risk"]["score"],
    }
    overall = fd["overall_score"]

    items = [
        {"id": "beneficiaries", "label": "We track the number of direct beneficiaries served",
         "dims": {"who": 0.6, "how_much": 0.3}, "sdgs": [1, 2, 3]},
        {"id": "outcomes", "label": "We measure outcomes (not just outputs) for beneficiaries",
         "dims": {"what": 0.7, "how_much": 0.4}, "sdgs": []},
        {"id": "ghg", "label": "We track greenhouse gas emissions or have reduction targets",
         "dims": {"risk": 0.5, "what": 0.3}, "sdgs": [13, 7]},
        {"id": "water", "label": "We measure water usage or have water stewardship practices",
         "dims": {"risk": 0.4, "what": 0.3}, "sdgs": [6, 14]},
        {"id": "gender", "label": "We track gender diversity in our workforce/beneficiaries",
         "dims": {"who": 0.5, "contribution": 0.2}, "sdgs": [5, 10]},
        {"id": "local_hiring", "label": "We prioritize hiring from local/underserved communities",
         "dims": {"contribution": 0.6, "who": 0.3}, "sdgs": [8, 10, 1]},
        {"id": "supply_chain", "label": "We have responsible supply chain policies",
         "dims": {"risk": 0.5, "contribution": 0.3}, "sdgs": [12, 8]},
        {"id": "baseline", "label": "We have baseline data from before our intervention started",
         "dims": {"contribution": 0.7, "how_much": 0.3}, "sdgs": []},
        {"id": "third_party", "label": "We have third-party verification or independent audits",
         "dims": {"risk": 0.8, "contribution": 0.4}, "sdgs": []},
        {"id": "theory_of_change", "label": "We have a documented Theory of Change",
         "dims": {"what": 0.5, "contribution": 0.4}, "sdgs": []},
        {"id": "stakeholder", "label": "We regularly collect feedback from beneficiaries/stakeholders",
         "dims": {"who": 0.4, "how_much": 0.3, "risk": 0.2}, "sdgs": []},
        {"id": "negative_screen", "label": "We assess and mitigate negative/unintended impacts",
         "dims": {"risk": 0.7}, "sdgs": []},
    ]

    import json as _json
    import re as _re
    items_json = _json.dumps(items)
    base_json = _json.dumps(base_scores)
    company_name_safe = _re.sub(r'[^a-zA-Z0-9]', '-', company_name).lower().strip('-') or 'default'

    return f"""
<h3 style="margin-top:24px">Improve Your Score</h3>
<p style="color:var(--text-secondary);font-size:0.85em;margin-bottom:12px">
Check practices your organization follows. The radar chart and scores above update in real-time.
</p>
<div id="interactive-panel" style="background:var(--surface);border-radius:var(--radius);box-shadow:var(--shadow-sm);padding:20px;border:1px solid var(--border);margin:0 0 16px">
<div id="checklist-items"></div>
</div>
<script>
(function() {{
  const items = {items_json};
  const baseScores = {base_json};
  const origOverall = {overall:.2f};
  const dimNames = ['what','who','how_much','contribution','risk'];
  const dimLabels = {{what:'What',who:'Who',how_much:'How Much',contribution:'Contribution',risk:'Risk'}};
  const dimKeys = {{what:'what',who:'who',how_much:'how_much',contribution:'contribution',risk:'risk'}};

  const container = document.getElementById('checklist-items');
  items.forEach(function(item) {{
    const div = document.createElement('div');
    div.style.cssText = 'padding:8px 0;border-bottom:1px solid #f0f0f0;display:flex;align-items:center;gap:12px';
    const cb = document.createElement('input');
    cb.type = 'checkbox'; cb.id = 'chk-'+item.id;
    cb.style.cssText = 'width:18px;height:18px;cursor:pointer;accent-color:#1976d2;flex-shrink:0';
    cb.addEventListener('change', recalc);
    const lbl = document.createElement('label');
    lbl.htmlFor = 'chk-'+item.id;
    lbl.textContent = item.label;
    lbl.style.cssText = 'cursor:pointer;font-size:0.9em;flex:1';
    const dims = Object.keys(item.dims).map(function(d) {{ return dimLabels[d]; }}).join(', ');
    const tag = document.createElement('span');
    tag.textContent = dims;
    tag.style.cssText = 'font-size:0.7em;color:var(--text-secondary);background:var(--primary-light);padding:2px 8px;border-radius:10px;white-space:nowrap';
    div.appendChild(cb); div.appendChild(lbl); div.appendChild(tag);
    container.appendChild(div);
  }});

  function getGrade(s) {{
    if (s >= 4.0) return 'A';
    if (s >= 3.0) return 'B';
    if (s >= 2.0) return 'C';
    if (s >= 1.0) return 'D';
    return 'F';
  }}
  function gradeClass(g) {{ return 'grade-'+g[0]; }}

  const STORAGE_KEY = 'impact-vision-scenario-' + '{company_name_safe}';

  function saveState() {{
    const state = {{}};
    items.forEach(function(item) {{
      const cb = document.getElementById('chk-'+item.id);
      if (cb) state[item.id] = cb.checked;
    }});
    try {{ localStorage.setItem(STORAGE_KEY, JSON.stringify(state)); }} catch(e) {{}}
  }}

  function loadState() {{
    try {{
      const saved = localStorage.getItem(STORAGE_KEY);
      if (!saved) return;
      const state = JSON.parse(saved);
      items.forEach(function(item) {{
        const cb = document.getElementById('chk-'+item.id);
        if (cb && state[item.id]) cb.checked = true;
      }});
      recalc();
    }} catch(e) {{}}
  }}

  function recalc() {{
    const scores = {{}};
    dimNames.forEach(function(d) {{ scores[d] = baseScores[d]; }});
    let anyChecked = false;
    items.forEach(function(item) {{
      const cb = document.getElementById('chk-'+item.id);
      if (cb && cb.checked) {{
        anyChecked = true;
        Object.keys(item.dims).forEach(function(d) {{
          scores[d] = Math.min(5.0, scores[d] + item.dims[d]);
        }});
      }}
    }});
    saveState();
    const avg = dimNames.reduce(function(s,d){{ return s+scores[d]; }}, 0) / 5;
    const grade = getGrade(avg);
    const delta = avg - origOverall;

    const gradeEl = document.getElementById('main-grade');
    if (gradeEl) {{
      gradeEl.textContent = grade;
      gradeEl.className = 'value ' + gradeClass(grade);
    }}
    const overallEl = document.getElementById('main-overall');
    if (overallEl) {{
      overallEl.innerHTML = avg.toFixed(1) + '<span style="font-size:0.5em;color:var(--text-secondary)">/5</span>';
    }}
    const deltaCard = document.getElementById('delta-card');
    const deltaEl = document.getElementById('main-delta');
    if (deltaCard && deltaEl) {{
      deltaCard.style.display = anyChecked ? 'block' : 'none';
      deltaEl.textContent = (delta >= 0 ? '+' : '') + delta.toFixed(1);
      deltaEl.style.color = delta > 0 ? 'var(--success)' : 'var(--text-secondary)';
    }}

    dimNames.forEach(function(d) {{
      const scoreEl = document.getElementById('dim-score-'+d);
      if (scoreEl) scoreEl.textContent = scores[d].toFixed(1) + '/5';
      const barEl = document.getElementById('dim-bar-'+d);
      if (barEl) {{
        const pct = Math.round(scores[d] / 5 * 100);
        barEl.style.width = pct + '%';
        barEl.className = 'bar-fill ' + (pct >= 60 ? 'green' : pct >= 30 ? 'orange' : 'red');
      }}
    }});

    const vals = dimNames.map(function(d) {{ return scores[d]; }});
    vals.push(vals[0]);
    const labels = dimNames.map(function(d) {{ return dimLabels[d]; }});
    labels.push(labels[0]);
    const baseVals = dimNames.map(function(d) {{ return baseScores[d]; }});
    baseVals.push(baseVals[0]);

    const traces = anyChecked ? [
      {{type:'scatterpolar', r:baseVals, theta:labels, fill:'toself', fillcolor:'rgba(176,190,197,0.15)',
        line:{{color:'#b0bec5',width:1.5,dash:'dot'}}, marker:{{size:5}}, name:'Before'}},
      {{type:'scatterpolar', r:vals, theta:labels, fill:'toself', fillcolor:'rgba(25,118,210,0.15)',
        line:{{color:'#1976d2',width:2.5}}, marker:{{size:7,color:'#1976d2'}}, name:'With improvements'}}
    ] : [
      {{type:'scatterpolar', r:vals, theta:labels, fill:'toself', fillcolor:'rgba(25,118,210,0.12)',
        line:{{color:'#1976d2',width:2.5}}, marker:{{size:7,color:'#1976d2'}}, name:'Current'}}
    ];

    Plotly.react('radar-chart', traces, {{
      polar:{{radialaxis:{{visible:true,range:[0,5],tickfont:{{size:10}}}},angularaxis:{{tickfont:{{size:11}}}}}},
      showlegend:true, legend:{{orientation:'h',y:-0.1}},
      height:380, margin:{{l:70,r:70,t:40,b:40}},
      paper_bgcolor:'transparent', plot_bgcolor:'transparent',
      font:{{family:'Inter, -apple-system, sans-serif'}}
    }}, {{responsive:true}});
  }}
  loadState();
}})();
</script>"""


def _generate_executive_summary(data: dict, company: dict) -> str:
    """Generate an executive summary section for the HTML report."""
    parts: list[str] = ['<div class="section"><h2>Executive Summary</h2><div class="content">']

    name = company.get("name", "the company")
    sector = company.get("sector", "")

    fd = data.get("five_dimensions")
    sdg = data.get("sdg_alignment")
    gw = data.get("greenwashing")

    if fd:
        parts.append(
            f"<p><strong>{name}</strong>"
            f"{f' ({sector})' if sector else ''} received an overall impact score of "
            f"<strong>{fd['overall_score']:.1f}/5.0 (Grade {fd['overall_grade']})</strong>. "
            f"Score confidence is rated as <em>{fd.get('overall_provenance', 'estimated')}</em>.</p>"
        )

    if sdg:
        top_sdgs = sorted(sdg, key=lambda s: s.get("score", 0), reverse=True)[:3]
        if top_sdgs:
            sdg_list = ", ".join(f"SDG {s['goal']} ({s.get('goal_name', '')})" for s in top_sdgs if s.get("score", 0) > 0)
            if sdg_list:
                parts.append(f"<p>Strongest SDG alignments: {sdg_list}.</p>")

    if gw:
        score = gw.get("overall_score", 0) if isinstance(gw, dict) else getattr(gw, "overall_score", 0)
        classification = gw.get("classification", "") if isinstance(gw, dict) else getattr(gw, "classification", "")
        if score:
            parts.append(f"<p>Greenwashing risk assessment: <strong>{classification}</strong> (score: {score}/100).</p>")

    gap = data.get("gap_analysis")
    if gap:
        coverage = gap.get("coverage_pct", 0)
        parts.append(f"<p>Core Metric Set coverage: {coverage}%.</p>")

    parts.append(
        "<p><em>This report was generated by Impact Vision using IRIS+ metrics, "
        "5 Dimensions of Impact scoring, and multi-framework ESG analysis. "
        "Scores based on limited data should be validated with additional evidence.</em></p>"
    )
    parts.append("</div></div>")
    return "\n".join(parts)


def _generate_report_narrative_prompt(data: dict) -> str:
    """Generate a structured prompt for LLM-based narrative generation.

    Returns a prompt that can be passed to an LLM to generate a polished
    executive summary and key findings narrative.
    """
    company = data.get("company", {})
    fd = data.get("five_dimensions")
    sdg = data.get("sdg_alignment")

    prompt = f"""Generate a professional impact assessment narrative for {company.get('name', 'this company')}.

Company: {company.get('name', 'Unknown')}
Sector: {company.get('sector', 'Not specified')}
Description: {company.get('description', 'N/A')[:500]}

"""
    if fd:
        prompt += f"""5-Dimension Score: {fd['overall_score']:.1f}/5.0 (Grade: {fd['overall_grade']})
Confidence: {fd.get('overall_provenance', 'estimated')}
"""
    if sdg:
        top = [s for s in sdg if s.get("score", 0) > 30][:5]
        if top:
            prompt += "Top SDG alignments: " + ", ".join(
                f"SDG {s['goal']} ({s.get('score', 0):.0f}%)" for s in top
            ) + "\n"

    prompt += """
Please write:
1. A 2-3 sentence executive summary
2. Key strengths (2-3 bullet points)
3. Key gaps or risks (2-3 bullet points)
4. Recommended next steps (2-3 items)

Use professional impact investment language. Be specific and evidence-based."""
    return prompt


def _metric_tracking_dashboard(data: dict) -> str:
    """Generate a metric tracking status dashboard (card grid)."""
    fd = data.get("five_dimensions", {})
    ga = data.get("gap_analysis", {})
    if not fd and not ga:
        return ""

    tracked: dict[str, dict] = {}
    gaps: dict[str, dict] = {}

    for dim_name in ("what", "who", "how_much", "contribution", "risk"):
        dim = fd.get(dim_name, {})
        for mt in dim.get("metrics_tracked", []):
            mid = mt if isinstance(mt, str) else str(mt)
            tracked[mid] = {"name": mid, "dimension": dim.get("dimension", dim_name), "status": "tracked"}
        for g in dim.get("gaps", []):
            gid = g.split(" (")[0] if isinstance(g, str) else str(g)
            if gid not in tracked:
                gaps[gid] = {"name": gid, "dimension": dim.get("dimension", dim_name), "status": "gap"}

    for m in ga.get("reported", []):
        mid = m.get("id", "")
        if mid and mid not in tracked:
            tracked[mid] = {"name": m.get("name", mid), "dimension": "", "status": "tracked"}
    for m in ga.get("missing", []):
        mid = m.get("id", "")
        if mid and mid not in tracked and mid not in gaps:
            gaps[mid] = {"name": m.get("name", mid), "dimension": "", "status": "gap"}

    all_metrics = list(tracked.values()) + list(gaps.values())
    if not all_metrics:
        return ""

    parts = [
        '<h2>Metric Tracking Dashboard</h2>',
        f'<p style="color:var(--text-secondary);font-size:0.88em;margin-bottom:12px">'
        f'{len(tracked)} tracked &bull; {len(gaps)} gaps &bull; {len(all_metrics)} total</p>',
        '<div class="metric-grid">',
    ]
    for m in sorted(all_metrics, key=lambda x: (x["status"] != "tracked", x["name"])):
        status = m["status"]
        status_label = "Tracked" if status == "tracked" else "Gap"
        parts.append(
            f'<div class="metric-card {status}">'
            f'<div class="mc-id">{m["name"]}</div>'
            f'<div class="mc-name">{m.get("dimension", "")}</div>'
            f'<span class="mc-status {status}">{status_label}</span>'
            f'</div>'
        )
    parts.append("</div>")
    return "\n".join(parts)


def _impact_claims_section(data: dict) -> str:
    """Generate expandable impact claim evidence cards."""
    claims = data.get("impact_claims", [])
    if not claims:
        return ""

    parts = [
        f'<h2>Impact Claims <span style="font-size:0.65em;color:var(--text-secondary);font-weight:400">'
        f'({len(claims)} claims)</span></h2>',
    ]

    visible = min(5, len(claims))
    for i, claim in enumerate(claims):
        hidden = ' style="display:none"' if i >= visible else ""
        cat = claim.get("category", "intent")
        conf = claim.get("confidence", 0.5)
        evidence = claim.get("evidence_strength", 1)
        text = claim.get("text", "No text")
        negated = claim.get("negated", False)
        metrics = claim.get("mapped_metrics", [])
        sdg_targets = claim.get("mapped_sdg_targets", [])
        conf_pct = int(conf * 100)
        conf_color = "#2e7d32" if conf >= 0.7 else "#f57c00" if conf >= 0.4 else "#c62828"

        parts.append(f'<div class="claim-card" data-claim-idx="{i}"{hidden}>')
        parts.append('<div class="claim-header" onclick="this.parentElement.querySelector(\'.claim-body\').classList.toggle(\'active\')">')
        parts.append(f'<span class="claim-badge {cat}">{cat}</span>')
        parts.append(f'<span class="claim-text">{text[:120]}{"..." if len(text) > 120 else ""}</span>')
        parts.append('<span class="claim-toggle">&#9660;</span>')
        parts.append("</div>")
        parts.append('<div class="claim-body">')

        if negated:
            parts.append('<div style="color:var(--danger);font-size:0.85em;margin-bottom:6px">&#9888; Negation detected in this claim</div>')

        parts.append('<div style="margin-bottom:8px"><strong style="font-size:0.82em">Confidence:</strong> ')
        parts.append(f'<div class="confidence-bar"><div class="confidence-fill" style="width:{conf_pct}%;background:{conf_color}"></div></div>')
        parts.append(f' <span style="font-size:0.82em;color:{conf_color}">{conf_pct}%</span></div>')

        parts.append('<div style="margin-bottom:8px"><strong style="font-size:0.82em">Evidence Level:</strong> ')
        stars = "&#9733;" * evidence + "&#9734;" * (5 - evidence)
        parts.append(f'<span style="color:#f9a825;font-size:0.95em">{stars}</span>')
        parts.append(f' <span style="font-size:0.78em;color:var(--text-secondary)">NESTA Level {evidence}</span></div>')

        if metrics:
            parts.append('<div style="margin-bottom:8px"><strong style="font-size:0.82em">Mapped Metrics:</strong> ')
            parts.append(" ".join(f'<span class="chip">{m}</span>' for m in metrics[:8]))
            parts.append("</div>")
        if sdg_targets:
            parts.append('<div style="margin-bottom:8px"><strong style="font-size:0.82em">SDG Targets:</strong> ')
            parts.append(" ".join(f'<span class="chip" style="background:#fff3e0;color:#e65100">{t}</span>' for t in sdg_targets[:8]))
            parts.append("</div>")

        parts.append("</div></div>")

    if len(claims) > visible:
        remaining = len(claims) - visible
        parts.append(
            f'<button id="show-more-claims" '
            f'style="display:block;margin:12px auto;padding:8px 24px;background:var(--primary-light);color:var(--primary);'
            f'border:1px solid var(--primary);border-radius:var(--radius-sm);cursor:pointer;font-size:0.88em">'
            f'Show {remaining} more claims</button>'
            '<script>'
            'document.getElementById("show-more-claims").addEventListener("click",function(){'
            'document.querySelectorAll(".claim-card").forEach(function(c){c.style.display=""});'
            'this.style.display="none"'
            '});'
            '</script>'
        )

    return "\n".join(parts)


def _impact_pathway_section(data: dict) -> str:
    """Generate an auto-inferred Theory of Change pathway diagram."""
    company = data.get("company", {})
    fd = data.get("five_dimensions", {})
    sdg = data.get("sdg_alignments", [])
    ga = data.get("gap_analysis", {})

    inputs_items = []
    if company.get("sector"):
        inputs_items.append(f"Sector: {company['sector']}")
    if company.get("geography"):
        inputs_items.append(f"Geography: {company['geography']}")
    reported_count = ga.get("metrics_reported", 0) if ga else 0
    if reported_count:
        inputs_items.append(f"{reported_count} metrics reported")
    if not inputs_items:
        inputs_items.append("Investment capital")

    activities = []
    if company.get("description"):
        desc = company["description"]
        if len(desc) > 60:
            desc = desc[:57] + "..."
        activities.append(desc)
    if company.get("impact_themes"):
        activities.extend(company["impact_themes"][:2])
    if not activities:
        activities.append("Core operations")

    outputs = []
    for dim_name in ("what", "who", "how_much"):
        dim = fd.get(dim_name, {})
        tracked = dim.get("metrics_tracked", [])
        if tracked:
            outputs.extend(str(t) for t in tracked[:2])
    if not outputs:
        outputs.append("Metrics to be reported")

    outcomes = []
    top_sdgs = sorted(sdg, key=lambda s: s.get("score", 0), reverse=True)[:3]
    for s in top_sdgs:
        if s.get("score", 0) > 0:
            outcomes.append(f"SDG {s['goal']}: {s.get('goal_name', '')[:25]}")
    if not outcomes:
        outcomes.append("SDG alignment pending")

    impact_items = []
    if fd:
        impact_items.append(f"5D Score: {fd.get('overall_score', 0):.1f}/5")
        impact_items.append(f"Grade: {fd.get('overall_grade', 'N/A')}")
    if not impact_items:
        impact_items.append("Impact TBD")

    stages = [
        ("Inputs", inputs_items),
        ("Activities", activities),
        ("Outputs", outputs),
        ("Outcomes", outcomes),
        ("Impact", impact_items),
    ]

    parts = ["<h2>Impact Pathway (Theory of Change)</h2>", '<div class="pathway">']
    for idx, (label, items) in enumerate(stages):
        items_html = "<br>".join(f"&bull; {it}" for it in items[:3])
        parts.append(
            f'<div class="pathway-stage">'
            f'<div class="pathway-box"><h4>{label}</h4>'
            f'<div class="items">{items_html}</div></div>'
        )
        if idx < len(stages) - 1:
            parts.append('<span class="pathway-arrow">&#8594;</span>')
        parts.append("</div>")
    parts.append("</div>")
    return "\n".join(parts)


def _to_html(data: dict) -> str:
    company = data["company"]
    sections: list[str] = []
    sdg_colors_map = {
        1: "#E5243B", 2: "#DDA63A", 3: "#4C9F38", 4: "#C5192D", 5: "#FF3A21",
        6: "#26BDE2", 7: "#FCC30B", 8: "#A21942", 9: "#FD6925", 10: "#DD1367",
        11: "#FD9D24", 12: "#BF8B2E", 13: "#3F7E44", 14: "#0A97D9", 15: "#56C02B",
        16: "#00689D", 17: "#19486A",
    }

    sections.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Impact Report: {company['name']}</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
:root {{
  --primary: #0d47a1; --primary-light: #e3f2fd; --primary-dark: #002171;
  --accent: #1976d2; --accent-light: #63a4ff;
  --success: #2e7d32; --success-light: #e8f5e9;
  --warning: #f57c00; --warning-light: #fff3e0;
  --danger: #c62828; --danger-light: #ffebee;
  --surface: #ffffff; --bg: #f5f7fa; --text: #1a1a2e; --text-secondary: #5f6368;
  --border: #e0e4e8; --shadow-sm: 0 1px 3px rgba(0,0,0,0.08); --shadow-md: 0 4px 12px rgba(0,0,0,0.1);
  --radius: 12px; --radius-sm: 8px;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 1080px; margin: 0 auto; padding: 32px 24px; color: var(--text); background: var(--bg); line-height: 1.5; }}
.report-header {{ background: linear-gradient(135deg, var(--primary) 0%, var(--accent) 100%); color: white; padding: 36px 40px; border-radius: var(--radius); margin-bottom: 28px; box-shadow: var(--shadow-md); }}
.report-header h1 {{ font-size: 1.75em; font-weight: 700; margin-bottom: 6px; letter-spacing: -0.02em; }}
.report-header .subtitle {{ opacity: 0.9; font-size: 0.95em; }}
.report-header .meta-row {{ display: flex; gap: 20px; flex-wrap: wrap; margin-top: 12px; opacity: 0.85; font-size: 0.85em; }}
.report-header .meta-row span {{ display: flex; align-items: center; gap: 4px; }}
.tag-row {{ display: flex; gap: 8px; flex-wrap: wrap; margin-top: 12px; }}
.tag {{ display: inline-block; background: rgba(255,255,255,0.2); padding: 3px 12px; border-radius: 20px; font-size: 0.8em; }}
h2 {{ color: var(--primary); font-size: 1.3em; margin: 32px 0 16px; padding-bottom: 8px; border-bottom: 2px solid var(--primary-light); font-weight: 600; }}
h3 {{ color: var(--text); font-size: 1.05em; margin: 20px 0 10px; font-weight: 600; }}
.cards-row {{ display: flex; gap: 16px; flex-wrap: wrap; margin: 16px 0; }}
.score-card {{ flex: 0 0 auto; background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow-sm); padding: 20px 28px; text-align: center; border: 1px solid var(--border); min-width: 130px; }}
.score-card .value {{ font-size: 2.2em; font-weight: 700; line-height: 1.1; }}
.score-card .label {{ font-size: 0.8em; color: var(--text-secondary); margin-top: 6px; text-transform: uppercase; letter-spacing: 0.03em; }}
.grade-A {{ color: var(--success); }} .grade-B {{ color: #558b2f; }} .grade-C {{ color: #f9a825; }} .grade-D {{ color: #e65100; }} .grade-F {{ color: var(--danger); }}
.chart-row {{ display: flex; gap: 20px; flex-wrap: wrap; margin: 16px 0; }}
.chart-box {{ flex: 1 1 420px; min-width: 0; background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow-sm); padding: 20px; border: 1px solid var(--border); overflow: hidden; }}
table {{ border-collapse: collapse; width: 100%; margin: 12px 0; background: var(--surface); border-radius: var(--radius-sm); overflow: hidden; box-shadow: var(--shadow-sm); font-size: 0.9em; }}
th {{ background: var(--primary); color: white; font-weight: 600; padding: 12px 14px; text-align: left; text-transform: uppercase; font-size: 0.75em; letter-spacing: 0.05em; }}
td {{ border-bottom: 1px solid var(--border); padding: 10px 14px; }}
tr:hover td {{ background: var(--primary-light); }}
.bar-track {{ background: #e8eaed; border-radius: 6px; height: 10px; width: 100%; overflow: hidden; }}
.bar-fill {{ height: 100%; border-radius: 6px; transition: width 0.4s ease; }}
.bar-fill.blue {{ background: linear-gradient(90deg, var(--accent), var(--accent-light)); }}
.bar-fill.green {{ background: linear-gradient(90deg, #43a047, #66bb6a); }}
.bar-fill.orange {{ background: linear-gradient(90deg, #ef6c00, #ffa726); }}
.bar-fill.red {{ background: linear-gradient(90deg, #c62828, #ef5350); }}
.bar-fill.coverage {{ background: linear-gradient(90deg, var(--primary), var(--accent)); }}
.rec {{ background: var(--warning-light); padding: 14px 18px; border-left: 4px solid var(--warning); margin: 8px 0; border-radius: 0 var(--radius-sm) var(--radius-sm) 0; font-size: 0.9em; line-height: 1.6; }}
.bm-delta.positive {{ color: var(--success); font-weight: 700; }}
.bm-delta.negative {{ color: var(--danger); font-weight: 700; }}
.bm-delta.neutral {{ color: var(--text-secondary); font-weight: 600; }}
.coverage-hero {{ background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow-sm); padding: 24px 28px; border: 1px solid var(--border); margin: 16px 0; display: flex; align-items: center; gap: 20px; }}
.coverage-hero .pct {{ font-size: 2.5em; font-weight: 800; color: var(--primary); line-height: 1; }}
.coverage-hero .detail {{ flex: 1; }}
.coverage-hero .bar-track {{ height: 14px; margin-top: 8px; }}
.footer {{ margin-top: 48px; padding: 20px 0; border-top: 2px solid var(--border); color: var(--text-secondary); font-size: 0.8em; text-align: center; }}
.footer a {{ color: var(--accent); text-decoration: none; }}
@media(max-width: 700px) {{ .chart-row {{ flex-direction: column; }} .chart-box {{ flex-basis: 100%; }} }}

/* 7.1.1 -- 5D Overlay Panel */
.dim-clickable {{ cursor: pointer; }}
.dim-clickable:hover td {{ background: var(--primary-light) !important; }}
.dim-overlay {{ display: none; }}
.dim-overlay.active {{ display: table-row; }}
.dim-overlay td {{ background: #f8fafe; padding: 16px 14px; }}
.dim-overlay-content {{ display: flex; gap: 16px; flex-wrap: wrap; }}
.dim-overlay-col {{ flex: 1 1 200px; min-width: 180px; }}
.dim-overlay-col h4 {{ font-size: 0.85em; color: var(--primary); margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.03em; }}
.metric-pill {{ display: inline-block; padding: 3px 10px; border-radius: 14px; font-size: 0.78em; margin: 2px 3px; font-weight: 500; }}
.metric-pill.tracked {{ background: var(--success-light); color: var(--success); }}
.metric-pill.gap {{ background: var(--danger-light); color: var(--danger); }}
.metric-pill.partial {{ background: var(--warning-light); color: var(--warning); }}
.dim-suggestion {{ font-size: 0.82em; color: var(--text-secondary); margin: 4px 0; padding-left: 12px; border-left: 2px solid var(--warning); }}

/* 7.1.2 -- SDG Drill-down */
.sdg-clickable {{ cursor: pointer; }}
.sdg-clickable:hover td {{ background: #fff9e6 !important; }}
.sdg-detail {{ display: none; }}
.sdg-detail.active {{ display: table-row; }}
.sdg-detail td {{ background: #fffef5; padding: 16px 14px; }}
.evidence-chain {{ display: flex; align-items: center; gap: 4px; flex-wrap: wrap; margin: 6px 0; }}
.chain-node {{ background: var(--primary-light); border: 1px solid var(--border); border-radius: var(--radius-sm); padding: 6px 10px; font-size: 0.78em; text-align: center; max-width: 160px; }}
.chain-arrow {{ color: var(--text-secondary); font-size: 1.1em; }}
.chain-confidence {{ font-size: 0.7em; color: var(--text-secondary); }}
.sdg-rec {{ font-size: 0.82em; padding: 6px 10px; background: var(--warning-light); border-radius: var(--radius-sm); margin: 3px 0; }}

/* 7.1.3 -- Metric Tracking Dashboard */
.metric-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 12px; margin: 16px 0; }}
.metric-card {{ background: var(--surface); border-radius: var(--radius-sm); box-shadow: var(--shadow-sm); padding: 14px 16px; border: 1px solid var(--border); border-top: 3px solid var(--border); }}
.metric-card.tracked {{ border-top-color: var(--success); }}
.metric-card.gap {{ border-top-color: var(--danger); }}
.metric-card.partial {{ border-top-color: var(--warning); }}
.metric-card .mc-id {{ font-size: 0.75em; color: var(--text-secondary); font-weight: 600; }}
.metric-card .mc-name {{ font-size: 0.88em; font-weight: 600; margin: 4px 0; }}
.metric-card .mc-status {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 0.72em; font-weight: 600; text-transform: uppercase; }}
.mc-status.tracked {{ background: var(--success-light); color: var(--success); }}
.mc-status.gap {{ background: var(--danger-light); color: var(--danger); }}
.mc-status.partial {{ background: var(--warning-light); color: var(--warning); }}

/* 7.1.4 -- Claim Evidence Cards */
.claim-card {{ background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow-sm); padding: 16px 20px; margin: 10px 0; border: 1px solid var(--border); transition: box-shadow 0.2s; }}
.claim-card:hover {{ box-shadow: var(--shadow-md); }}
.claim-header {{ display: flex; align-items: center; gap: 10px; cursor: pointer; }}
.claim-text {{ flex: 1; font-size: 0.9em; }}
.claim-badge {{ display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 0.72em; font-weight: 600; text-transform: uppercase; }}
.claim-badge.outcome {{ background: #e8f5e9; color: #2e7d32; }}
.claim-badge.output {{ background: #e3f2fd; color: #1565c0; }}
.claim-badge.activity {{ background: #fff3e0; color: #e65100; }}
.claim-badge.intent {{ background: #f3e5f5; color: #7b1fa2; }}
.claim-badge.risk {{ background: #ffebee; color: #c62828; }}
.claim-body {{ display: none; margin-top: 12px; padding-top: 12px; border-top: 1px solid var(--border); }}
.claim-body.active {{ display: block; }}
.confidence-bar {{ background: #e8eaed; border-radius: 6px; height: 8px; width: 100px; display: inline-block; overflow: hidden; vertical-align: middle; }}
.confidence-fill {{ height: 100%; border-radius: 6px; }}
.chip {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 0.75em; margin: 2px; background: var(--primary-light); color: var(--primary); }}
.claim-toggle {{ color: var(--text-secondary); font-size: 0.85em; }}

/* 7.2.3 -- Impact Pathway Diagram */
.pathway {{ display: flex; align-items: stretch; gap: 0; margin: 20px 0; overflow-x: auto; padding: 10px 0; }}
.pathway-stage {{ flex: 1 1 150px; min-width: 130px; text-align: center; position: relative; }}
.pathway-box {{ background: var(--surface); border: 2px solid var(--primary-light); border-radius: var(--radius-sm); padding: 12px 10px; margin: 0 6px; min-height: 80px; display: flex; flex-direction: column; justify-content: center; }}
.pathway-box h4 {{ font-size: 0.78em; color: var(--primary); text-transform: uppercase; margin-bottom: 6px; }}
.pathway-box .items {{ font-size: 0.78em; color: var(--text); }}
.pathway-box .confidence-label {{ font-size: 0.68em; color: var(--text-secondary); margin-top: 4px; }}
.pathway-arrow {{ position: absolute; right: -10px; top: 50%; transform: translateY(-50%); color: var(--primary); font-size: 1.4em; z-index: 1; }}

/* 7.1.5 -- Print/PDF */
@media print {{
  body {{ background: white; max-width: 100%; padding: 16px; font-size: 10pt; }}
  .report-header {{ background: #0d47a1 !important; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
  .score-card, .chart-box, .claim-card, .metric-card {{ break-inside: avoid; }}
  #interactive-panel, .dim-clickable .claim-toggle {{ display: none !important; }}
  .dim-overlay, .sdg-detail, .claim-body {{ display: table-row !important; }}
  .claim-body {{ display: block !important; }}
  h2 {{ break-before: page; }}
  .footer {{ break-before: avoid; }}
}}

/* 7.1.6 -- Report Comparison */
.comparison-table {{ width: 100%; border-collapse: collapse; }}
.comparison-table th {{ background: var(--primary); color: white; padding: 10px; font-size: 0.8em; }}
.comparison-table td {{ padding: 8px 10px; border-bottom: 1px solid var(--border); font-size: 0.88em; }}
.delta-up {{ color: var(--success); font-weight: 700; }}
.delta-down {{ color: var(--danger); font-weight: 700; }}
.delta-same {{ color: var(--text-secondary); }}
</style>
</head>
<body>
<div class="report-header">
<h1>Impact Assessment Report</h1>
<p class="subtitle">{company['name']}{' | ' + company.get('sector', '') if company.get('sector') else ''}</p>
<div class="meta-row">
  <span>Generated: {data['generated_at'][:10]}</span>
  <span>Standard: {data['catalog_version']}</span>
</div>""")

    if company.get("impact_themes") or company.get("sdg_claims"):
        sections.append('<div class="tag-row">')
        for t in company.get("impact_themes", []):
            sections.append(f'<span class="tag">{t}</span>')
        for g in company.get("sdg_claims", []):
            sections.append(f'<span class="tag">SDG {g}</span>')
        sections.append('</div>')
    sections.append('</div>')

    sections.append(_generate_executive_summary(data, company))

    comp_html = _comparison_section(data)
    if comp_html:
        sections.append(comp_html)

    if "five_dimensions" in data:
        fd = data["five_dimensions"]
        grade_class = f"grade-{fd['overall_grade'][0]}"

        sections.append(f"""
<h2>5 Dimensions of Impact</h2>
<div class="cards-row">
<div class="score-card"><div class="value {grade_class}" id="main-grade">{fd['overall_grade']}</div><div class="label">Overall Grade</div></div>
<div class="score-card"><div class="value" id="main-overall">{fd['overall_score']:.1f}<span style="font-size:0.5em;color:var(--text-secondary)">/5</span></div><div class="label">Overall Score</div></div>
<div class="score-card"><div class="value" style="font-size:0.9em;color:{'var(--success)' if fd.get('overall_provenance') == 'evidence-based' else 'var(--warning)' if fd.get('overall_provenance') == 'partial' else 'var(--danger)'}">{'Evidence-Based' if fd.get('overall_provenance') == 'evidence-based' else 'Partial' if fd.get('overall_provenance') == 'partial' else 'Estimated'}</div><div class="label">Confidence</div></div>
<div class="score-card" id="delta-card" style="display:none"><div class="value" id="main-delta" style="font-size:1.4em;color:var(--text-secondary)">+0.0</div><div class="label">Improvement</div></div>
</div>
<div class="chart-row">
<div class="chart-box" id="radar-chart"></div>
<div class="chart-box">
<table id="dim-table">
<tr><th>Dimension</th><th>Score</th><th style="min-width:160px">Progress</th><th>Confidence</th><th>Metrics</th><th>Details</th></tr>
""")
        dims_js = []
        scores_js = []
        for dim_name in ["what", "who", "how_much", "contribution", "risk"]:
            dim = fd[dim_name]
            pct = int(dim["score"] / 5.0 * 100)
            bar_color = "green" if pct >= 60 else "orange" if pct >= 30 else "red"
            dims_js.append(f'"{dim["dimension"]}"')
            scores_js.append(str(dim["score"]))
            reported = dim.get("metrics_reported", 0)
            available = dim.get("metrics_available", 0)
            metric_pct = int(reported / available * 100) if available > 0 else 0
            metric_color = "var(--success)" if metric_pct >= 50 else "var(--warning)" if metric_pct > 0 else "var(--danger)"
            gaps_preview = ", ".join(g.split(" (")[0] for g in dim.get("gaps", [])[:3])
            gaps_tooltip = f' title="{gaps_preview}"' if gaps_preview else ""
            prov = dim.get("provenance", "estimated")
            prov_color = "var(--success)" if prov == "evidence-based" else "var(--warning)" if prov == "partial" else "var(--text-secondary)"
            prov_label = prov.replace("-", " ").title() if prov else "Estimated"
            sections.append(f"""<tr class="dim-clickable" data-dim="{dim_name}">
<td><strong>{dim['dimension']}</strong> <span style="font-size:0.7em;color:var(--text-secondary)">&#9660;</span></td>
<td style="font-weight:600" id="dim-score-{dim_name}">{dim['score']}/5</td>
<td><div class="bar-track"><div class="bar-fill {bar_color}" id="dim-bar-{dim_name}" style="width:{pct}%"></div></div></td>
<td style="font-size:0.8em;color:{prov_color};font-weight:500">{prov_label}</td>
<td style="font-size:0.85em;white-space:nowrap"><span style="color:{metric_color};font-weight:600">{reported}/{available}</span>
<span style="color:var(--text-secondary)"{gaps_tooltip}>{' tracked' if reported > 0 else ' not tracked'}</span></td>
<td style="font-size:0.85em;color:var(--text-secondary)">{dim['notes']}</td>
</tr>""")

            tracked_metrics = dim.get("metrics_tracked", [])
            gap_metrics = dim.get("gaps", [])
            recs = dim.get("recommendations", fd.get("recommendations", []))[:3]
            sections.append(f'<tr class="dim-overlay" id="overlay-{dim_name}"><td colspan="6"><div class="dim-overlay-content">')
            sections.append('<div class="dim-overlay-col"><h4>Tracked Metrics</h4>')
            if tracked_metrics:
                for tm in tracked_metrics[:6]:
                    label = tm if isinstance(tm, str) else str(tm)
                    sections.append(f'<span class="metric-pill tracked">{label}</span>')
            else:
                sections.append(f'<span style="font-size:0.82em;color:var(--text-secondary)">No metrics tracked ({available} available)</span>')
            sections.append('</div><div class="dim-overlay-col"><h4>Gaps</h4>')
            if gap_metrics:
                for gm in gap_metrics[:6]:
                    label = gm.split(" (")[0] if isinstance(gm, str) else str(gm)
                    sections.append(f'<span class="metric-pill gap">{label}</span>')
            else:
                sections.append('<span style="font-size:0.82em;color:var(--success)">No gaps identified</span>')
            sections.append('</div><div class="dim-overlay-col"><h4>Suggestions</h4>')
            if recs:
                for rc in recs[:3]:
                    sections.append(f'<div class="dim-suggestion">{rc}</div>')
            else:
                sections.append('<span style="font-size:0.82em;color:var(--text-secondary)">Report more metrics to receive suggestions</span>')
            sections.append('</div></div></td></tr>')

        sections.append("</table>")
        sections.append("""<script>
document.querySelectorAll('.dim-clickable').forEach(function(row) {
  row.addEventListener('click', function() {
    var dim = this.dataset.dim;
    var overlay = document.getElementById('overlay-' + dim);
    if (overlay) overlay.classList.toggle('active');
  });
});
</script>""")

        overall_prov = fd.get("overall_provenance", "estimated")
        if overall_prov in ("estimated", "partial"):
            disclaimer_color = "var(--warning-light)" if overall_prov == "partial" else "var(--danger-light)"
            disclaimer_border = "var(--warning)" if overall_prov == "partial" else "var(--danger)"
            disclaimer_text = (
                "Scores are primarily <strong>estimated</strong> from keywords and sector heuristics. "
                "Report at least 3 IRIS+ metrics per dimension for evidence-based scores."
            ) if overall_prov == "estimated" else (
                "Some scores are based on <strong>partial evidence</strong>. "
                "Additional reported metrics will strengthen the assessment."
            )
            sections.append(f"""
<div style="margin:12px 0;padding:12px 16px;background:{disclaimer_color};border-left:4px solid {disclaimer_border};border-radius:var(--radius-sm);font-size:0.85em">
{disclaimer_text}
</div>""")

        total_reported = sum(fd[d].get("metrics_reported", 0) for d in ["what", "who", "how_much", "contribution", "risk"])
        total_available = sum(fd[d].get("metrics_available", 0) for d in ["what", "who", "how_much", "contribution", "risk"])
        if total_available > 0:
            sections.append(f"""
<div style="margin-top:12px;padding:12px 16px;background:var(--primary-light);border-radius:var(--radius-sm);font-size:0.85em">
<strong>Metric Tracking:</strong> {total_reported} of {total_available} available IRIS+ metrics reported
({int(total_reported / total_available * 100)}% coverage).
{f'Top gaps: {", ".join(g.split(" (")[0] for g in fd["what"].get("gaps", [])[:3] + fd["who"].get("gaps", [])[:2])}' if total_reported < total_available else 'Full coverage achieved!'}
</div>""")
        sections.append("</div></div>")

        sections.append(f"""<script>
Plotly.newPlot('radar-chart', [{{
  type: 'scatterpolar', r: [{','.join(scores_js)},{scores_js[0]}],
  theta: [{','.join(dims_js)},{dims_js[0]}],
  fill: 'toself', fillcolor: 'rgba(25,118,210,0.12)',
  line: {{color: '#1976d2', width: 2.5}}, marker: {{size: 7, color: '#1976d2'}},
  name: 'Current'
}}], {{
  polar: {{radialaxis: {{visible: true, range: [0, 5], tickfont: {{size: 10}}}}, angularaxis: {{tickfont: {{size: 11}}}}}},
  showlegend: true, legend: {{orientation: 'h', y: -0.1}},
  height: 380, margin: {{l: 70, r: 70, t: 40, b: 40}},
  paper_bgcolor: 'transparent', plot_bgcolor: 'transparent',
  font: {{family: 'Inter, -apple-system, sans-serif'}}
}}, {{responsive: true}});
</script>""")

        sections.append(_interactive_scoring_section(fd, data.get("sdg_alignments", []), company.get("name", "")))

        if fd.get("recommendations"):
            sections.append("<h3>Recommendations</h3>")
            for r in fd["recommendations"]:
                sections.append(f'<div class="rec">{r}</div>')

    if "target_tracking" in data:
        tt = data["target_tracking"]
        targets = tt.get("targets", [])
        if targets:
            sections.append("<h2>Impact Target Tracking</h2>")
            sections.append("<table><tr><th>Metric</th><th>Target</th><th>Current</th><th>Progress</th><th>Status</th></tr>")
            status_colors = {
                "exceeded": "var(--success)", "on_track": "var(--success)",
                "behind": "var(--warning)", "at_risk": "var(--danger)", "no_data": "var(--text-secondary)",
            }
            status_icons = {
                "exceeded": "🟢", "on_track": "🟢",
                "behind": "🟡", "at_risk": "🔴", "no_data": "⚪",
            }
            for t in targets:
                color = status_colors.get(t["status"], "var(--text-secondary)")
                icon = status_icons.get(t["status"], "")
                pct = t.get("progress_pct", 0)
                pct_display = f"{pct:.0f}%" if pct else "N/A"
                sections.append(f"""<tr>
<td>{t['metric_id']}</td>
<td>{t.get('target_description', 'N/A')}</td>
<td>{t.get('current_value', 'N/A')}</td>
<td><div class="bar-track"><div class="bar-fill" style="width:{min(pct, 100):.0f}%;background:{color}"></div></div> {pct_display}</td>
<td style="color:{color};font-weight:600">{icon} {t['status'].replace('_', ' ').title()}</td>
</tr>""")
            sections.append("</table>")
            summary = tt.get("summary", {})
            if summary:
                sections.append(f"""<div style="margin-top:12px;padding:12px 16px;background:var(--primary-light);border-radius:var(--radius-sm);font-size:0.85em">
<strong>Target Summary:</strong> {summary.get('on_track', 0)} on track, {summary.get('behind', 0)} behind, {summary.get('exceeded', 0)} exceeded, {summary.get('at_risk', 0)} at risk
</div>""")

    if "beneficiary_feedback" in data:
        bf = data["beneficiary_feedback"]
        sections.append("<h2>Beneficiary Feedback</h2>")
        sections.append('<div class="chart-row">')
        if bf.get("satisfaction_score") is not None:
            sat = bf["satisfaction_score"]
            sat_color = "var(--success)" if sat >= 4 else "var(--warning)" if sat >= 3 else "var(--danger)"
            sections.append(f'<div class="score-card"><div class="value" style="color:{sat_color}">{sat}/5</div><div class="label">Satisfaction</div></div>')
        if bf.get("nps") is not None:
            nps = bf["nps"]
            nps_color = "var(--success)" if nps >= 50 else "var(--warning)" if nps >= 0 else "var(--danger)"
            sections.append(f'<div class="score-card"><div class="value" style="color:{nps_color}">{nps}</div><div class="label">NPS</div></div>')
        if bf.get("sample_size"):
            sections.append(f'<div class="score-card"><div class="value">{bf["sample_size"]}</div><div class="label">Sample Size</div></div>')
        if bf.get("quality_of_life_improvement") is not None:
            sections.append(f'<div class="score-card"><div class="value" style="color:var(--success)">{bf["quality_of_life_improvement"]}%</div><div class="label">QoL Improvement</div></div>')
        if bf.get("would_recommend") is not None:
            sections.append(f'<div class="score-card"><div class="value">{bf["would_recommend"]}%</div><div class="label">Would Recommend</div></div>')
        sections.append('</div>')
        if bf.get("methodology"):
            sections.append(f'<p style="font-size:0.85em;color:var(--text-secondary)">Methodology: {bf["methodology"]}')
            if bf.get("survey_date"):
                sections[-1] += f' | Survey date: {bf["survey_date"]}'
            sections[-1] += '</p>'
        if bf.get("themes"):
            sections.append('<div style="margin-top:8px"><strong>Positive Themes:</strong> ')
            sections.append(', '.join(f'<span style="background:var(--primary-light);padding:2px 8px;border-radius:12px;font-size:0.85em">{t}</span>' for t in bf["themes"][:5]))
            sections.append('</div>')
        if bf.get("challenges"):
            sections.append('<div style="margin-top:8px"><strong>Challenges:</strong> ')
            sections.append(', '.join(f'<span style="background:#fff3e0;padding:2px 8px;border-radius:12px;font-size:0.85em">{c}</span>' for c in bf["challenges"][:5]))
            sections.append('</div>')
        if bf.get("quotes"):
            sections.append('<div style="margin-top:12px">')
            for q in bf["quotes"][:3]:
                sections.append(f'<blockquote style="border-left:3px solid var(--primary);padding:8px 16px;margin:8px 0;font-style:italic;color:var(--text-secondary)">&ldquo;{q}&rdquo;</blockquote>')
            sections.append('</div>')

    if "sdg_alignments" in data:
        aligned = [a for a in data["sdg_alignments"] if a["score"] > 0]
        sdg_labels = [f'"SDG {a["goal"]}"' for a in aligned]
        sdg_scores = [str(a["score"]) for a in aligned]
        sdg_colors = []
        for a in aligned:
            sdg_colors.append(f'"{sdg_colors_map.get(a["goal"], "#1976d2")}"')

        sections.append(f"""
<h2>SDG Alignment <span style="font-size:0.65em;color:var(--text-secondary);font-weight:400">({len(aligned)} goals)</span></h2>
<div class="chart-row">
<div class="chart-box" id="sdg-chart"></div>
<div class="chart-box">
<table id="sdg-table">
<tr><th>SDG</th><th>Name</th><th>Score</th><th>Confidence</th><th>Matched Metrics</th></tr>
""")
        for a in aligned:
            sdg_color = sdg_colors_map.get(a["goal"], "#666")
            metrics_str = ", ".join(a.get("matched_metrics", [])[:3])
            conf_style = {
                "high": "color:var(--success);font-weight:600",
                "medium": "color:#f57c00;font-weight:600",
                "low": "color:var(--danger);font-weight:600",
            }.get(a["confidence"], "")
            sections.append(f"""<tr class="sdg-clickable" data-sdg="{a['goal']}">
<td><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{sdg_color};margin-right:6px;vertical-align:middle"></span>SDG {a['goal']} <span style="font-size:0.7em;color:var(--text-secondary)">&#9660;</span></td>
<td>{a.get('goal_name', '')}</td>
<td style="font-weight:600">{a['score']}</td><td style="{conf_style}">{a['confidence']}</td><td style="font-size:0.85em">{metrics_str}</td>
</tr>""")

            sections.append(f'<tr class="sdg-detail" id="sdg-detail-{a["goal"]}"><td colspan="5">')
            targets = a.get("matched_targets", [])
            if targets:
                sections.append('<div style="margin-bottom:8px"><strong style="font-size:0.82em">Targets:</strong> ')
                sections.append(" ".join(f'<span class="chip">{t}</span>' for t in targets))
                sections.append("</div>")
            chains = a.get("evidence_chain", [])
            if chains:
                sections.append('<div style="margin-bottom:8px"><strong style="font-size:0.82em">Evidence Chains:</strong>')
                for ch in chains[:5]:
                    claim_text = str(ch.get("claim_text", ""))[:60]
                    metric = ch.get("metric_id", "")
                    ev_type = str(ch.get("evidence_type", "")).replace("_", " ").title()
                    sdg_tgt = ch.get("sdg_target", "")
                    conf = ch.get("confidence", 0)
                    sections.append('<div class="evidence-chain">')
                    if claim_text:
                        sections.append(f'<span class="chain-node">{claim_text}</span><span class="chain-arrow">&#8594;</span>')
                    if metric:
                        sections.append(f'<span class="chain-node">{metric}</span><span class="chain-arrow">&#8594;</span>')
                    sections.append(f'<span class="chain-node">{ev_type}</span><span class="chain-arrow">&#8594;</span>')
                    sections.append(f'<span class="chain-node">{sdg_tgt}<br><span class="chain-confidence">{conf:.0%} conf.</span></span>')
                    sections.append('</div>')
                sections.append("</div>")
            sdg_recs = a.get("recommendations", [])
            if sdg_recs:
                sections.append('<div><strong style="font-size:0.82em">Recommendations:</strong>')
                for sr in sdg_recs[:3]:
                    sections.append(f'<div class="sdg-rec">{sr}</div>')
                sections.append("</div>")
            if not targets and not chains and not sdg_recs:
                sections.append(f'<span style="font-size:0.85em;color:var(--text-secondary)">Provenance: {a.get("provenance", "estimated")}. Report more metrics to build evidence chain.</span>')
            sections.append("</td></tr>")

        sections.append("</table></div></div>")
        sections.append("""<script>
document.querySelectorAll('.sdg-clickable').forEach(function(row) {
  row.addEventListener('click', function() {
    var sdg = this.dataset.sdg;
    var detail = document.getElementById('sdg-detail-' + sdg);
    if (detail) detail.classList.toggle('active');
  });
});
</script>""")

        if sdg_labels:
            sections.append(f"""<script>
Plotly.newPlot('sdg-chart', [{{
  type: 'bar', x: [{','.join(sdg_labels)}], y: [{','.join(sdg_scores)}],
  marker: {{color: [{','.join(sdg_colors)}], line: {{width: 0}}, cornerradius: 4}},
  text: [{','.join(sdg_scores)}], textposition: 'outside', textfont: {{size: 11, family: 'Inter, sans-serif'}}
}}], {{
  yaxis: {{range: [0, 110], title: 'Alignment Score', gridcolor: '#f0f0f0', titlefont: {{size: 12}}}},
  xaxis: {{tickangle: -30, tickfont: {{size: 10}}}},
  height: 380, margin: {{l: 55, r: 20, t: 20, b: 60}},
  paper_bgcolor: 'transparent', plot_bgcolor: 'transparent',
  font: {{family: 'Inter, -apple-system, sans-serif'}}
}}, {{responsive: true}});
</script>""")

    claims_html = _impact_claims_section(data)
    if claims_html:
        sections.append(claims_html)

    if "impact_analysis" in data:
        ia = data["impact_analysis"]
        sections.append("""
<h2>Impact Opportunities & Risks</h2>
<div class="chart-row">
<div class="chart-box">
<h3 style="color:var(--success)">Opportunities</h3>""")
        for o in ia.get("opportunities", []):
            sections.append(f'<div class="rec" style="border-left-color:var(--success);background:var(--success-light)">+ {o}</div>')
        sections.append("""</div>
<div class="chart-box">
<h3 style="color:var(--danger)">Risks</h3>""")
        for r in ia.get("risks", []):
            sections.append(f'<div class="rec" style="border-left-color:var(--danger);background:var(--danger-light)">! {r}</div>')
        sections.append("</div></div>")

    if "gap_analysis" in data:
        ga = data["gap_analysis"]
        pct = ga["coverage_percentage"]
        pct_color = "var(--success)" if pct >= 60 else "var(--warning)" if pct >= 30 else "var(--danger)"
        bar_cls = "green" if pct >= 60 else "orange" if pct >= 30 else "red"
        sections.append(f"""
<h2>Gap Analysis</h2>
<div class="coverage-hero">
  <div class="pct" style="color:{pct_color}">{ga['coverage_percentage']}%</div>
  <div class="detail">
    <div style="font-weight:600;font-size:1.1em">Core Metric Set Coverage</div>
    <div style="color:var(--text-secondary);font-size:0.9em">{ga['metrics_reported']} of {ga['core_metric_set_size']} metrics reported | {ga['metrics_missing']} missing</div>
    <div class="bar-track"><div class="bar-fill {bar_cls}" style="width:{int(pct)}%"></div></div>
  </div>
</div>
""")
        if ga.get("missing"):
            sections.append("""<h3>Missing Metrics</h3><table><tr><th>ID</th><th>Name</th><th>Definition</th></tr>""")
            for m in ga["missing"]:
                sections.append(f"<tr><td style='font-weight:600'>{m['id']}</td><td>{m['name']}</td><td style='font-size:0.85em;color:var(--text-secondary)'>{m.get('definition', '')[:120]}</td></tr>")
            sections.append("</table>")

        if ga.get("recommendations"):
            sections.append("<h3>Recommendations</h3>")
            for r in ga["recommendations"]:
                sections.append(f'<div class="rec">{r}</div>')

    dashboard_html = _metric_tracking_dashboard(data)
    if dashboard_html:
        sections.append(dashboard_html)

    if "greenwashing" in data:
        gw = data["greenwashing"]
        gw_score = gw.get("overall_score", 0)
        gw_class = gw.get("classification", "Unknown")
        gw_color = "#2e7d32" if gw_score < 30 else "#f57c00" if gw_score < 60 else "#c62828"
        sections.append(f"""
<h2>Greenwashing / Impact-Washing Risk</h2>
<div class="cards-row">
<div class="score-card" style="border-left:4px solid {gw_color}">
  <div class="value" style="color:{gw_color}">{gw_score}</div>
  <div class="label">Risk Score (0-100)</div>
</div>
<div class="score-card"><div class="value" style="font-size:1.2em;color:{gw_color}">{gw_class}</div><div class="label">Classification</div></div>
</div>""")
        sub = gw.get("sub_scores", {})
        if sub:
            sections.append('<table><tr><th>Sub-Score</th><th>Value</th></tr>')
            for sname, sval in sub.items():
                display = sname.replace("_", " ").title()
                sections.append(f'<tr><td>{display}</td><td>{sval}/100</td></tr>')
            sections.append('</table>')
        if gw.get("flags"):
            sections.append('<h3>Flags</h3>')
            for flag in gw["flags"]:
                sections.append(f'<div class="rec" style="border-left-color:{gw_color}">{flag}</div>')
        if gw.get("recommendations"):
            sections.append('<h3>Recommendations</h3>')
            for rec in gw["recommendations"]:
                sections.append(f'<div class="rec">{rec}</div>')

    if "benchmark_comparison" in data:
        bm = data["benchmark_comparison"]
        sections.append(f"""
<h2>Sector Benchmark Comparison</h2>
<p style="color:var(--text-secondary);font-size:0.9em;margin-bottom:12px">{bm['sector']} | {bm.get('sample_note', '')}</p>
<div class="chart-row">
<div class="chart-box" id="benchmark-chart"></div>
<div class="chart-box">
<table>
<tr><th>Dimension</th><th>Your Score</th><th>Benchmark</th><th>Delta</th></tr>
""")
        ov = bm["overall"]
        delta_class = "positive" if ov["delta"] > 0 else ("negative" if ov["delta"] < 0 else "neutral")
        delta_arrow = "+" if ov["delta"] > 0 else ""
        sections.append(f"""<tr style="font-weight:600;background:var(--primary-light)">
<td>Overall</td><td>{ov['actual']:.1f}</td><td>{ov['benchmark']:.1f}</td>
<td class="bm-delta {delta_class}">{delta_arrow}{ov['delta']:.1f}</td></tr>""")

        bm_dim_labels = []
        bm_actual_vals = []
        bm_benchmark_vals = []
        for dim, vals in bm["dimensions"].items():
            delta_class = "positive" if vals["delta"] > 0 else ("negative" if vals["delta"] < 0 else "neutral")
            delta_arrow = "+" if vals["delta"] > 0 else ""
            display_name = dim.replace('_', ' ').title()
            bm_dim_labels.append(f'"{display_name}"')
            bm_actual_vals.append(str(vals["actual"]))
            bm_benchmark_vals.append(str(vals["benchmark"]))
            sections.append(f"""<tr>
<td>{display_name}</td><td>{vals['actual']:.1f}</td><td>{vals['benchmark']:.1f}</td>
<td class="bm-delta {delta_class}">{delta_arrow}{vals['delta']:.1f}</td></tr>""")
        sections.append("</table></div></div>")

        if bm_dim_labels:
            sections.append(f"""<script>
Plotly.newPlot('benchmark-chart', [
  {{type:'bar', name:'Your Score', x:[{','.join(bm_dim_labels)}], y:[{','.join(bm_actual_vals)}], marker:{{color:'#1976d2', cornerradius:3}}}},
  {{type:'bar', name:'Benchmark', x:[{','.join(bm_dim_labels)}], y:[{','.join(bm_benchmark_vals)}], marker:{{color:'#b0bec5', cornerradius:3}}}}
], {{
  barmode:'group', yaxis:{{range:[0,5.5], title:'Score', gridcolor:'#f0f0f0'}},
  height:350, margin:{{l:50,r:20,t:20,b:50}}, legend:{{orientation:'h', y:1.08}},
  paper_bgcolor:'transparent', plot_bgcolor:'transparent',
  font:{{family:'Inter, -apple-system, sans-serif', size:11}}
}}, {{responsive:true}});
</script>""")

    pathway_html = _impact_pathway_section(data)
    if pathway_html:
        sections.append(pathway_html)

    sections.append("""
<div class="footer">
Generated by <a href="#">Impact Vision</a> &mdash; Open-source impact measurement engine &mdash; IRIS+ 5.3c
</div>
</body></html>""")
    return "\n".join(sections)


def _to_pdf(html: str, output_path: str, context) -> ToolResult:
    """Convert HTML to PDF using WeasyPrint if available, otherwise provide instructions."""
    if output_path:
        path = Path(output_path)
        if not path.is_absolute():
            path = context.cwd / path
        if not path.suffix:
            path = path.with_suffix(".pdf")
    else:
        path = context.cwd / "impact_report.pdf"

    try:
        import weasyprint  # type: ignore[import-untyped]
        weasyprint.HTML(string=html).write_pdf(str(path))
        return ToolResult(
            output=f"PDF report saved to: {path}",
            metadata={"output_path": str(path), "format": "pdf"},
        )
    except ImportError:
        html_path = path.with_suffix(".html")
        html_path.parent.mkdir(parents=True, exist_ok=True)
        html_path.write_text(html, encoding="utf-8")
        return ToolResult(
            output=(
                f"WeasyPrint not installed. HTML saved to: {html_path}\n"
                "To generate PDF:\n"
                "  Option 1: pip install weasyprint  (then re-run)\n"
                "  Option 2: Open the HTML file in a browser and use Ctrl+P / Cmd+P to print to PDF\n"
                "The report includes print-friendly CSS for clean PDF output."
            ),
            metadata={"output_path": str(html_path), "format": "html_for_pdf"},
        )


def _load_comparison_data(assessment_id: str, current: dict) -> dict:
    """Load a previous assessment and compute deltas against the current report."""
    try:
        from openharness.impact.storage import AssessmentStore
        store = AssessmentStore()
        prev = store.get_assessment(assessment_id)
    except Exception:
        return {"error": f"Could not load assessment {assessment_id}"}

    if prev is None:
        return {"error": f"Assessment {assessment_id} not found"}

    prev_data = prev if isinstance(prev, dict) else prev.model_dump()
    comparison: dict = {"previous_id": assessment_id, "dimensions": {}, "sdg": {}}

    cur_fd = current.get("five_dimensions", {})
    prev_fd = prev_data.get("five_dimensions", {})
    if cur_fd and prev_fd:
        for dim_name in ("what", "who", "how_much", "contribution", "risk"):
            cur_score = cur_fd.get(dim_name, {}).get("score", 0)
            prev_score = prev_fd.get(dim_name, {}).get("score", 0)
            comparison["dimensions"][dim_name] = {
                "current": cur_score, "previous": prev_score,
                "delta": round(cur_score - prev_score, 2),
            }
        comparison["overall"] = {
            "current": cur_fd.get("overall_score", 0),
            "previous": prev_fd.get("overall_score", 0),
            "delta": round(cur_fd.get("overall_score", 0) - prev_fd.get("overall_score", 0), 2),
        }

    cur_sdg = {a["goal"]: a for a in current.get("sdg_alignments", [])}
    prev_sdg = {a["goal"]: a for a in prev_data.get("sdg_alignments", [])}
    for goal in sorted(set(cur_sdg) | set(prev_sdg)):
        c = cur_sdg.get(goal, {}).get("score", 0)
        p = prev_sdg.get(goal, {}).get("score", 0)
        if c or p:
            comparison["sdg"][goal] = {"current": c, "previous": p, "delta": round(c - p, 2)}

    return comparison


def _comparison_section(data: dict) -> str:
    """Render comparison table if comparison data is present."""
    comp = data.get("comparison", {})
    if not comp or comp.get("error"):
        return ""

    parts = [
        f'<h2>Assessment Comparison <span style="font-size:0.65em;color:var(--text-secondary);font-weight:400">'
        f'vs. {comp.get("previous_id", "previous")}</span></h2>',
    ]

    if comp.get("overall"):
        ov = comp["overall"]
        delta_cls = "delta-up" if ov["delta"] > 0 else "delta-down" if ov["delta"] < 0 else "delta-same"
        parts.append(
            f'<div class="cards-row">'
            f'<div class="score-card"><div class="value">{ov["current"]:.1f}</div><div class="label">Current</div></div>'
            f'<div class="score-card"><div class="value" style="color:var(--text-secondary)">{ov["previous"]:.1f}</div><div class="label">Previous</div></div>'
            f'<div class="score-card"><div class="value {delta_cls}">{"+" if ov["delta"]>0 else ""}{ov["delta"]:.1f}</div><div class="label">Change</div></div>'
            f'</div>'
        )

    dims = comp.get("dimensions", {})
    if dims:
        parts.append('<table class="comparison-table"><tr><th>Dimension</th><th>Previous</th><th>Current</th><th>Change</th></tr>')
        for dim, vals in dims.items():
            display = dim.replace("_", " ").title()
            delta = vals["delta"]
            cls = "delta-up" if delta > 0 else "delta-down" if delta < 0 else "delta-same"
            arrow = "&#9650;" if delta > 0 else "&#9660;" if delta < 0 else "&#8212;"
            parts.append(
                f'<tr><td><strong>{display}</strong></td>'
                f'<td>{vals["previous"]:.1f}</td><td>{vals["current"]:.1f}</td>'
                f'<td class="{cls}">{arrow} {"+" if delta>0 else ""}{delta:.1f}</td></tr>'
            )
        parts.append("</table>")

    sdg_comp = comp.get("sdg", {})
    if sdg_comp:
        parts.append('<h3>SDG Score Changes</h3>')
        parts.append('<table class="comparison-table"><tr><th>SDG</th><th>Previous</th><th>Current</th><th>Change</th></tr>')
        for goal, vals in sorted(sdg_comp.items()):
            delta = vals["delta"]
            cls = "delta-up" if delta > 0 else "delta-down" if delta < 0 else "delta-same"
            arrow = "&#9650;" if delta > 0 else "&#9660;" if delta < 0 else "&#8212;"
            parts.append(
                f'<tr><td>SDG {goal}</td>'
                f'<td>{vals["previous"]:.0f}</td><td>{vals["current"]:.0f}</td>'
                f'<td class="{cls}">{arrow} {"+" if delta>0 else ""}{delta:.0f}</td></tr>'
            )
        parts.append("</table>")

    return "\n".join(parts)


def _to_target_progress_text(data: dict) -> str:
    """Generate a focused target progress report with trajectory projections."""
    company = data.get("company", {})
    lines = [
        "=" * 60,
        f"TARGET PROGRESS REPORT: {company.get('name', 'Unknown')}",
        f"Generated: {data.get('generated_at', '')}",
        "=" * 60, "",
    ]
    tt = data.get("target_tracking", {})
    targets = tt.get("targets", [])
    if not targets:
        lines.append("No impact targets defined. Set targets to enable progress tracking.")
        return "\n".join(lines)

    summary = tt.get("summary", {})
    lines.append(f"Total Targets: {len(targets)}")
    lines.append(f"  On Track: {summary.get('on_track', 0)} | Behind: {summary.get('behind', 0)}")
    lines.append(f"  Exceeded: {summary.get('exceeded', 0)} | At Risk: {summary.get('at_risk', 0)}")
    lines.append("")

    for t in targets:
        status = t.get("status", "no_data")
        pct = t.get("progress_pct", 0)
        lines.append(f"  {t['metric_id']}: {status.replace('_', ' ').upper()}")
        lines.append(f"    Target: {t.get('target_description', 'N/A')}")
        lines.append(f"    Current: {t.get('current_value', 'N/A')} | Progress: {pct:.0f}%")
        if pct > 0 and pct < 100:
            remaining = 100 - pct
            if remaining > 50:
                lines.append(f"    Trajectory: At risk — {remaining:.0f}% remaining, consider intervention")
            elif remaining > 20:
                lines.append(f"    Trajectory: On track — {remaining:.0f}% remaining")
            else:
                lines.append(f"    Trajectory: Near completion — {remaining:.0f}% remaining")
        lines.append("")

    if data.get("five_dimensions"):
        fd = data["five_dimensions"]
        lines.append(f"5D Score Context: {fd.get('overall_score', 0):.1f}/5 (Grade: {fd.get('overall_grade', 'N/A')})")

    return "\n".join(lines)


def _to_lp_ready_text(data: dict) -> str:
    """Generate an LP-formatted individual company report."""
    company = data.get("company", {})
    fd = data.get("five_dimensions", {})
    sdg = data.get("sdg_alignments", [])
    ga = data.get("gap_analysis", {})
    gw = data.get("greenwashing", {})

    lines = [
        "=" * 70,
        "CONFIDENTIAL — FOR LP DISTRIBUTION",
        "=" * 70,
        f"Company: {company.get('name', 'Unknown')}",
        f"Sector: {company.get('sector', 'N/A')}",
        f"Geography: {company.get('geography', 'N/A')}",
        f"Assessment Date: {data.get('generated_at', '')[:10]}",
        f"Standard: {data.get('catalog_version', 'IRIS+ 5.3c')}",
        "", "─" * 70, "",
        "EXECUTIVE SUMMARY",
        "─" * 40,
    ]

    if fd:
        lines.append(f"Overall Impact Score: {fd.get('overall_score', 0):.1f}/5.0 (Grade: {fd.get('overall_grade', 'N/A')})")
        lines.append(f"Assessment Confidence: {fd.get('overall_provenance', 'estimated').replace('-', ' ').title()}")

    top_sdgs = sorted(sdg, key=lambda s: s.get("score", 0), reverse=True)[:3]
    if top_sdgs:
        lines.append("Top SDG Alignments: " + ", ".join(
            f"SDG {s['goal']} ({s.get('score', 0):.0f}%)" for s in top_sdgs if s.get("score", 0) > 0
        ))

    if ga:
        lines.append(f"Core Metric Coverage: {ga.get('coverage_percentage', 0)}%")

    if isinstance(gw, dict) and gw.get("classification"):
        lines.append(f"Greenwashing Risk: {gw.get('classification', 'N/A')} ({gw.get('overall_score', 0)}/100)")

    lines.extend(["", "─" * 70, "", "IMPACT DIMENSIONS", "─" * 40])
    if fd:
        for dim_name in ("what", "who", "how_much", "contribution", "risk"):
            dim = fd.get(dim_name, {})
            lines.append(f"  {dim.get('dimension', dim_name)}: {dim.get('score', 0)}/5 | {dim.get('notes', '')}")

    lines.extend(["", "KEY RISKS AND RECOMMENDATIONS", "─" * 40])
    if fd and fd.get("recommendations"):
        for r in fd["recommendations"][:5]:
            lines.append(f"  • {r}")

    ia = data.get("impact_analysis", {})
    if ia.get("risks"):
        for r in ia["risks"][:3]:
            lines.append(f"  ⚠ {r}")

    lines.extend([
        "", "─" * 70,
        "This assessment was generated by Impact Vision using IRIS+ metrics,",
        "5 Dimensions of Impact scoring, and multi-framework ESG analysis.",
        "All scores should be validated with additional evidence and due diligence.",
        "─" * 70,
    ])
    return "\n".join(lines)


def _to_xlsx(data: dict, output_path: str, context) -> ToolResult:
    """Generate an Excel workbook with the impact report."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return ToolResult(output="openpyxl required for XLSX. Install: pip install openpyxl", is_error=True)

    if not output_path:
        return ToolResult(output="output_path is required for xlsx format", is_error=True)

    wb = Workbook()
    company = data["company"]

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")

    ws.append(["Impact Assessment Report"])
    ws.append(["Company", company["name"]])
    ws.append(["Sector", company.get("sector", "")])
    ws.append(["Generated", data["generated_at"]])
    ws.append(["Standard", data["catalog_version"]])
    ws.append([])

    if "five_dimensions" in data:
        fd = data["five_dimensions"]
        ws.append(["5 DIMENSIONS OF IMPACT"])
        headers = ["Dimension", "Score", "Metrics Reported", "Available", "Notes"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for dim_name in ["what", "who", "how_much", "contribution", "risk"]:
            dim = fd[dim_name]
            ws.append([dim["dimension"], dim["score"], dim["metrics_reported"], dim["metrics_available"], dim["notes"]])
        ws.append(["Overall", fd["overall_score"], "", "", fd["overall_grade"]])
        ws.append([])

    # Sheet 2: SDG Alignment
    if "sdg_alignments" in data:
        ws2 = wb.create_sheet("SDG Alignment")
        headers = ["SDG Goal", "Name", "Score", "Confidence", "Matched Metrics", "Matched Targets"]
        ws2.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for a in data["sdg_alignments"]:
            if a["score"] > 0:
                ws2.append([
                    f"SDG {a['goal']}", a.get("goal_name", ""), a["score"],
                    a["confidence"], ", ".join(a.get("matched_metrics", [])[:5]),
                    ", ".join(a.get("matched_targets", [])[:5]),
                ])

    # Sheet 3: Gap Analysis
    if "gap_analysis" in data:
        ws3 = wb.create_sheet("Gap Analysis")
        ga = data["gap_analysis"]
        ws3.append(["Coverage %", ga["coverage_percentage"]])
        ws3.append(["Reported", ga["metrics_reported"]])
        ws3.append(["Missing", ga["metrics_missing"]])
        ws3.append([])
        headers = ["Status", "Metric ID", "Name", "Value"]
        ws3.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws3.cell(row=5, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for m in ga.get("reported", []):
            ws3.append(["Reported", m["id"], m["name"], m.get("value", "")])
        for m in ga.get("missing", []):
            ws3.append(["Missing", m["id"], m["name"], ""])

    path = Path(output_path)
    if not path.is_absolute():
        path = context.cwd / path
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))

    return ToolResult(
        output=f"XLSX report saved to: {path}\nCompany: {company['name']}\nSheets: Summary, SDG Alignment, Gap Analysis",
        metadata={"output_path": str(path), "format": "xlsx"},
    )
