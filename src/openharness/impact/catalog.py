"""IRIS+ 5.3c Catalog ETL: parse the 263-column Excel into normalized JSON."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import openpyxl

from openharness.impact.models import DimensionTags, JointImpactIndicators, Metric

logger = logging.getLogger(__name__)

CATALOG_SHEET = "IRIS+ Catalog of Metrics (5.3c)"
GLOSSARY_SHEET = "Glossary"
REFERENCE_LISTS_SHEET = "Reference Lists"

# Column indices (0-based) from the IRIS+ 5.3c Excel
COL_ID = 0
COL_NAME = 1
COL_DEFINITION = 2
COL_FOOTNOTE = 3
COL_CALCULATION = 4
COL_USAGE_GUIDANCE = 5
COL_PRIMARY_CATEGORY = 6
COL_CROSS_CATEGORY = 7
# Columns 9-53: Impact Theme boolean flags
THEME_COLS_START = 9
THEME_COLS_END = 53
COL_ENVIRONMENTAL = 55
COL_SOCIAL = 56
COL_SECTION = 57
COL_SUBSECTION = 58
COL_CITATION = 59
COL_METRIC_TYPE = 60
COL_RELATED = 61
COL_METRIC_LEVEL = 62
COL_QUANTITY_TYPE = 63
COL_REPORTING_FORMAT = 64
# Column 65: SDGs summary (often empty; real data in individual target cols)
# Columns 66-232: Individual SDG Goal + Target boolean flags
SDG_COLS_START = 66
SDG_COLS_END = 232
# Columns 234-241: Dimensions of Impact
COL_DIM_WHAT = 234
COL_DIM_WHO = 235
COL_DIM_SCALE = 236
COL_DIM_DEPTH = 237
COL_DIM_DURATION = 238
COL_DIM_CONTRIB_DEPTH = 239
COL_DIM_CONTRIB_DUR = 240
COL_DIM_RISK = 241
# JII columns
COL_JII_GENDER = 243
COL_JII_JOBS = 244
COL_JII_CLIMATE = 245
# Stakeholder columns
COL_STAKE_CLIENTS = 249
COL_STAKE_DISTRIBUTORS = 250
COL_STAKE_EMPLOYEES = 251
COL_STAKE_ENVIRONMENT = 252
COL_STAKE_SUPPLIERS = 253
# Financial columns
COL_FIN_BALANCE = 255
COL_FIN_CASHFLOW = 256
COL_FIN_INCOME = 257
COL_FIN_OTHER = 258


def _str_or_none(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _str_val(val: Any) -> str:
    return _str_or_none(val) or ""


def _is_marked(val: Any) -> bool:
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ("X", "TRUE", "1", "YES")


def _parse_sdg_columns(row: tuple, header: tuple) -> tuple[list[int], list[str]]:
    """Extract SDG goals and targets from the wide boolean columns."""
    goals: set[int] = set()
    targets: list[str] = []

    for col_idx in range(SDG_COLS_START, min(SDG_COLS_END + 1, len(row))):
        if not _is_marked(row[col_idx]):
            continue
        col_name = _str_val(header[col_idx]) if col_idx < len(header) else ""
        if not col_name:
            continue

        if col_name.startswith("SDG Target "):
            target_id = col_name.replace("SDG Target ", "").strip().rstrip(".")
            if target_id:
                targets.append(target_id)
                try:
                    goal_num = int(target_id.split(".")[0])
                    goals.add(goal_num)
                except ValueError:
                    pass
        elif col_name.startswith("SDG ") and ":" in col_name:
            try:
                goal_num = int(col_name.split(":")[0].replace("SDG ", "").strip())
                goals.add(goal_num)
            except ValueError:
                pass

    return sorted(goals), sorted(targets)


def _parse_theme_columns(row: tuple, header: tuple) -> list[str]:
    """Extract active impact themes from boolean columns 9-53."""
    themes: list[str] = []
    for col_idx in range(THEME_COLS_START, min(THEME_COLS_END + 1, len(row))):
        if _is_marked(row[col_idx]):
            theme_name = _str_val(header[col_idx]) if col_idx < len(header) else ""
            if theme_name:
                themes.append(theme_name)
    return themes


def _determine_focus(row: tuple) -> str:
    env = _is_marked(row[COL_ENVIRONMENTAL]) if COL_ENVIRONMENTAL < len(row) else False
    soc = _is_marked(row[COL_SOCIAL]) if COL_SOCIAL < len(row) else False
    if env and soc:
        return "both"
    if env:
        return "environmental"
    if soc:
        return "social"
    return "both"


def _parse_related_metrics(val: Any) -> list[str]:
    if val is None:
        return []
    raw = str(val).strip()
    if not raw:
        return []
    parts = [p.strip() for p in raw.replace(";", ",").split(",")]
    return [p for p in parts if p]


def _parse_stakeholders(row: tuple) -> list[str]:
    mapping = {
        COL_STAKE_CLIENTS: "Clients",
        COL_STAKE_DISTRIBUTORS: "Distributors",
        COL_STAKE_EMPLOYEES: "Employees",
        COL_STAKE_ENVIRONMENT: "Environment",
        COL_STAKE_SUPPLIERS: "Suppliers",
    }
    return [label for col, label in mapping.items() if col < len(row) and _is_marked(row[col])]


def _parse_financials(row: tuple) -> list[str]:
    mapping = {
        COL_FIN_BALANCE: "Balance Sheet",
        COL_FIN_CASHFLOW: "Cash Flow",
        COL_FIN_INCOME: "Income Statement",
        COL_FIN_OTHER: "Other Financial",
    }
    return [label for col, label in mapping.items() if col < len(row) and _is_marked(row[col])]


def parse_metric_row(row: tuple, header: tuple) -> Metric | None:
    """Parse a single row from the IRIS+ catalog into a Metric object."""
    metric_id = _str_or_none(row[COL_ID])
    if not metric_id:
        return None

    name = _str_val(row[COL_NAME])
    if not name:
        return None

    sdg_goals, sdg_targets = _parse_sdg_columns(row, header)
    impact_themes = _parse_theme_columns(row, header)

    metric_type_raw = _str_val(row[COL_METRIC_TYPE]).lower() if COL_METRIC_TYPE < len(row) else ""
    metric_type = "submetric" if "sub" in metric_type_raw else "metric"

    dimensions = DimensionTags(
        what=_is_marked(row[COL_DIM_WHAT]) if COL_DIM_WHAT < len(row) else False,
        who=_is_marked(row[COL_DIM_WHO]) if COL_DIM_WHO < len(row) else False,
        how_much_scale=_is_marked(row[COL_DIM_SCALE]) if COL_DIM_SCALE < len(row) else False,
        how_much_depth=_is_marked(row[COL_DIM_DEPTH]) if COL_DIM_DEPTH < len(row) else False,
        how_much_duration=_is_marked(row[COL_DIM_DURATION]) if COL_DIM_DURATION < len(row) else False,
        contribution_depth=_is_marked(row[COL_DIM_CONTRIB_DEPTH]) if COL_DIM_CONTRIB_DEPTH < len(row) else False,
        contribution_duration=_is_marked(row[COL_DIM_CONTRIB_DUR]) if COL_DIM_CONTRIB_DUR < len(row) else False,
        risk=_is_marked(row[COL_DIM_RISK]) if COL_DIM_RISK < len(row) else False,
    )

    jii = JointImpactIndicators(
        gender=_is_marked(row[COL_JII_GENDER]) if COL_JII_GENDER < len(row) else False,
        jobs=_is_marked(row[COL_JII_JOBS]) if COL_JII_JOBS < len(row) else False,
        climate=_is_marked(row[COL_JII_CLIMATE]) if COL_JII_CLIMATE < len(row) else False,
    )

    return Metric(
        id=metric_id,
        name=name,
        definition=_str_val(row[COL_DEFINITION]),
        footnote=_str_or_none(row[COL_FOOTNOTE]) if COL_FOOTNOTE < len(row) else None,
        calculation=_str_or_none(row[COL_CALCULATION]) if COL_CALCULATION < len(row) else None,
        usage_guidance=_str_or_none(row[COL_USAGE_GUIDANCE]) if COL_USAGE_GUIDANCE < len(row) else None,
        primary_impact_category=_str_val(row[COL_PRIMARY_CATEGORY]),
        is_cross_category=_is_marked(row[COL_CROSS_CATEGORY]) if COL_CROSS_CATEGORY < len(row) else False,
        impact_themes=impact_themes,
        focus=_determine_focus(row),
        section=_str_val(row[COL_SECTION]) if COL_SECTION < len(row) else "",
        subsection=_str_val(row[COL_SUBSECTION]) if COL_SUBSECTION < len(row) else "",
        citation=_str_val(row[COL_CITATION]) if COL_CITATION < len(row) else "",
        metric_type=metric_type,
        related_metrics=_parse_related_metrics(row[COL_RELATED] if COL_RELATED < len(row) else None),
        metric_level=_str_val(row[COL_METRIC_LEVEL]) if COL_METRIC_LEVEL < len(row) else "",
        quantity_type=_str_val(row[COL_QUANTITY_TYPE]) if COL_QUANTITY_TYPE < len(row) else "",
        reporting_format=_str_val(row[COL_REPORTING_FORMAT]) if COL_REPORTING_FORMAT < len(row) else "",
        sdg_goals=sdg_goals,
        sdg_targets=sdg_targets,
        dimensions=dimensions,
        jii=jii,
        stakeholders=_parse_stakeholders(row),
        financials=_parse_financials(row),
    )


def load_catalog_from_excel(excel_path: str | Path) -> list[Metric]:
    """Load all metrics from the IRIS+ 5.3c Excel catalog."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"IRIS+ catalog not found: {path}")

    logger.info("Loading IRIS+ catalog from %s", path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    if CATALOG_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{CATALOG_SHEET}' not found in {path}")

    ws = wb[CATALOG_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return []

    header = rows[0]
    metrics: list[Metric] = []
    for row in rows[1:]:
        metric = parse_metric_row(row, header)
        if metric is not None:
            metrics.append(metric)

    logger.info("Loaded %d metrics from IRIS+ catalog", len(metrics))
    return metrics


def load_glossary_from_excel(excel_path: str | Path) -> dict[str, str]:
    """Load glossary terms from the IRIS+ catalog Excel."""
    path = Path(excel_path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if GLOSSARY_SHEET not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[GLOSSARY_SHEET]
    glossary: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        term = _str_or_none(row[0]) if len(row) > 0 else None
        defn = _str_or_none(row[1]) if len(row) > 1 else None
        if term and defn:
            glossary[term] = defn
    wb.close()
    return glossary


def save_catalog_json(metrics: list[Metric], output_path: str | Path) -> None:
    """Save processed catalog to JSON."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    data = [m.model_dump() for m in metrics]
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Saved %d metrics to %s", len(data), path)


def load_catalog_json(json_path: str | Path) -> list[Metric]:
    """Load processed catalog from JSON."""
    path = Path(json_path)
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    return [Metric.model_validate(item) for item in data]


def get_default_excel_path() -> Path:
    """Return the default path to the IRIS+ catalog Excel."""
    candidates = [
        Path(__file__).parent.parent.parent.parent / "data" / "raw" / "IRIS 5.3c Catalog of Metrics.xlsx",
        Path.cwd() / "data" / "raw" / "IRIS 5.3c Catalog of Metrics.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


def get_default_json_path() -> Path:
    """Return the default path for the processed JSON catalog."""
    candidates = [
        Path(__file__).parent.parent.parent.parent / "data" / "processed" / "iris_catalog_5.3c.json",
        Path.cwd() / "data" / "processed" / "iris_catalog_5.3c.json",
    ]
    for path in candidates:
        if path.parent.exists():
            return path
    return candidates[0]
